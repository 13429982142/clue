# -*- coding: utf-8 -*-
"""
商务线索挖掘与转化智能体 (PRD v1.0)
FastAPI 后端实现
"""
import os
import uuid
import json
import logging
import time
import re
import io
from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime, timezone, timedelta
from decimal import Decimal
from urllib.parse import urljoin, urlparse, parse_qs, parse_qsl, urlencode, urlunparse, unquote, quote

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))

from pydantic import BaseModel, Field
from fastapi import FastAPI, HTTPException, Depends, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy import create_engine, Column, String, Integer, Numeric, Text, Date, DateTime
from sqlalchemy.orm import sessionmaker, Session, declarative_base
import requests
from bs4 import BeautifulSoup

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
)
logger = logging.getLogger("lead_backend")
crawler_logger = logging.getLogger("lead_backend.crawler")

DEFAULT_SEARCH_CONFIG = {
    "monitorSources": [],
    "keywords": [],
    "excludeKeywords": [],
    "regions": [],
    "industries": [],
    "categories": [],
    "frequency": "每天早上8:00实时",
    "budgetMin": 0,
    "budgetMax": 2000,
    "purchaseTypes": [],
    "timeRange": "不限",
}

DEFAULT_CRAWLER_KEYWORDS = ["污水厂智能加药", "污水处理智能化改造"]
DEFAULT_CRAWLER_REGIONS = ["广东省"]
DEFAULT_CRAWLER_PURCHASE_TYPES = ["国企", "事业单位", "民企"]
DEFAULT_CRAWLER_SOURCES = ["全国公共资源交易平台"]
DEFAULT_CRAWLER_INDUSTRIES = ["污水自控", "智慧工厂", "高危污染监控"]
DEFAULT_CRAWLER_CATEGORIES = ["机电/节能改造", "环保在线监视"]
WATER_DOMAIN_TERMS = ["污水", "污水处理", "污水管网", "雨污", "雨污管网", "排水管网", "一体化污水处理"]
KEYWORD_SYNONYM_GROUPS = [
    WATER_DOMAIN_TERMS,
    ["改造", "提升改造", "系统改造", "管网改造", "缺陷修复", "延伸项目"],
    ["智能化", "自动化", "在线监测", "智慧水务", "少人化"],
    ["加药", "智能加药", "精准加药", "药剂投加"],
    ["曝气", "鼓风曝气", "曝气控制"],
]
DEFAULT_HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; LeadCrawler/1.0; +https://example.local)",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.6",
}
SOURCE_URL_MAP = {
    "中国政府采购网": ["https://www.ccgp.gov.cn/cggg/dfgg/gkzb/"],
    "全国公共资源交易平台": ["https://www.ggzy.gov.cn/deal/dealList.html"],
    "湖北省公共资源交易云平台": [
        "https://www.hbbidcloud.cn/hubei/search/fullsearch.html?wd=%E6%B1%A1%E6%B0%B4&cnum=001",
    ],
}
CRAWLER_REQUEST_TIMEOUT = float(os.getenv("CRAWLER_REQUEST_TIMEOUT", "10"))
CRAWLER_MAX_LINKS_PER_SOURCE = int(os.getenv("CRAWLER_MAX_LINKS_PER_SOURCE", "30"))
CRAWLER_MAX_PAGES = int(os.getenv("CRAWLER_MAX_PAGES", "5"))  # 列表页最多翻几页
CRAWLER_ALLOW_INSECURE_SSL = os.getenv("CRAWLER_ALLOW_INSECURE_SSL", "false").lower() == "true"
CRAWLER_ALLOW_HTTP_FALLBACK = os.getenv("CRAWLER_ALLOW_HTTP_FALLBACK", "true").lower() == "true"
SOURCE_FETCH_MAX_ATTEMPTS = 3  # 1 次初始 + 2 次当场重试
SOURCE_FETCH_RETRY_DELAY_SEC = float(os.getenv("SOURCE_FETCH_RETRY_DELAY_SEC", "2"))
# 常见列表分页查询参数（兼容多数招投标站点）
LIST_PAGE_QUERY_KEYS = (
    "page", "pageNo", "pageNum", "pagenum", "pn", "p",
    "currentPage", "current", "pageIndex", "index", "Page", "PAGE",
)
# 湖北公共资源云平台全文检索：一次返回全量结果，前端再切片分页
HBBID_FULLSEARCH_API = os.getenv(
    "HBBID_FULLSEARCH_API",
    "https://www.hbbidcloud.cn/inteligentsearch/rest/esinteligentsearch/getFullTextDataNew",
)
HBBID_FULLSEARCH_BATCH_RN = int(os.getenv("HBBID_FULLSEARCH_BATCH_RN", "500"))  # 站点前端上限 500
HBBID_FULLSEARCH_BASE = "https://www.hbbidcloud.cn"
# 搜索框型源：按监控关键词逐词查询并去重合并（单次 wd 组合召回不稳定）
HBBID_MAX_KEYWORD_QUERIES = int(os.getenv("HBBID_MAX_KEYWORD_QUERIES", "8"))

# --- 预调研：大模型（可选，混合引擎） ---
# 配置 LLM_API_KEY 后自动切换为大模型生成；未配置时使用规则引擎。
# 兼容 OpenAI 风格 /chat/completions 接口（如 OpenAI / 通义 / DeepSeek / 自建网关）。
LLM_API_KEY = os.getenv("LLM_API_KEY", "").strip()
LLM_API_BASE = os.getenv("LLM_API_BASE", "https://api.deepseek.com/v1").rstrip("/")
LLM_MODEL = os.getenv("LLM_MODEL", "deepseek-v4-flash")
LLM_TIMEOUT = float(os.getenv("LLM_TIMEOUT", "60"))

# --- DeepSeek 联网搜索（可选，开启后自动补充企查/行业新闻/环保公示等） ---
# 开启后，DeepSeek 生成预调研报告时会自动搜索互联网补充信息。
# 无需额外申请 API Key，直接使用 LLM_API_KEY，消耗 tokens 约多 20%。
ENABLE_DEEPSEEK_SEARCH = os.getenv("ENABLE_DEEPSEEK_SEARCH", "false").strip().lower() == "true"

# 我方业务能力库（用于痛点匹配与切入点建议）
BUSINESS_CAPABILITIES = [
    {
        "name": "智能加药",
        "keywords": ["加药", "投加", "药剂", "PAC", "PAM", "除磷", "脱氮", "carbon", "碳源"],
        "value": "基于水质在线监测的精准加药控制，平均节省药剂 15%-30%，稳定出水指标。",
    },
    {
        "name": "智能曝气",
        "keywords": ["曝气", "鼓风", "溶解氧", "DO", "风机", "能耗", "电耗"],
        "value": "DO 闭环精准曝气，降低鼓风机电耗 10%-25%，是吨水电耗的主要优化点。",
    },
    {
        "name": "少人化运维",
        "keywords": ["运维", "少人化", "无人值守", "人工", "人员", "巡检", "托管", "运营"],
        "value": "集中监控 + 智能巡检，减少现场值守人力，降低人工成本与人为操作风险。",
    },
    {
        "name": "综合智能化改造",
        "keywords": ["改造", "自动化", "智能化", "智慧水务", "在线监测", "数字化", "信息化"],
        "value": "PLC/SCADA 与数据中台一体化改造，打通全厂数据，支撑工艺优化与决策。",
    },
]

# 污水行业共性痛点（用于无明确信号时的行业先验）
WATER_INDUSTRY_PAIN_POINTS = [
    "吨水电耗偏高，曝气/水泵能耗占运行成本比重大。",
    "药剂投加多依赖人工经验，存在过量投加与出水波动风险。",
    "运维人员配置偏多、老龄化，少人化/无人值守改造需求上升。",
    "在线仪表与自控系统老旧，数据孤岛严重，难以支撑精细化运营。",
    "出水标准趋严（一级A/准Ⅳ类等），稳定达标压力大。",
]

# 聚合信息的目标字段（缺失时如实标记，禁止编造）
RESEARCH_INFO_FIELDS = [
    ("companyBasic", "企业基本信息"),
    ("plantScale", "厂区规模"),
    ("treatmentProcess", "处理工艺"),
    ("pastProjects", "过往改造项目"),
    ("operationMode", "运维模式"),
    ("energyLaborCost", "能耗与人工成本"),
]

# 数据库连接
DATABASE_URL = os.getenv("DATABASE_URL", "")

# MinIO 文件存储
MINIO_ENDPOINT = os.getenv("MINIO_ENDPOINT", "")
MINIO_ACCESS_KEY = os.getenv("MINIO_ACCESS_KEY", "")
MINIO_SECRET_KEY = os.getenv("MINIO_SECRET_KEY", "")
MINIO_BUCKET_NAME = os.getenv("MINIO_BUCKET_NAME", "bsrsaq")

# MinIO 客户端（懒初始化）
_minio_client = None

def _get_minio_client():
    global _minio_client
    if _minio_client is not None:
        return _minio_client
    if not MINIO_ENDPOINT or not MINIO_ACCESS_KEY or not MINIO_SECRET_KEY:
        return None
    try:
        from minio import Minio
        # 去掉协议前缀，Minio SDK 自行处理
        endpoint = MINIO_ENDPOINT.replace("http://", "").replace("https://", "")
        secure = MINIO_ENDPOINT.startswith("https://")
        _minio_client = Minio(
            endpoint,
            access_key=MINIO_ACCESS_KEY,
            secret_key=MINIO_SECRET_KEY,
            secure=secure,
        )
        # 确保桶存在
        if not _minio_client.bucket_exists(MINIO_BUCKET_NAME):
            _minio_client.make_bucket(MINIO_BUCKET_NAME)
    except Exception as exc:
        import logging
        logging.getLogger("lead_backend").warning("MinIO 初始化失败: %s", exc)
        _minio_client = None
    return _minio_client

engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

app = FastAPI(title="商务线索挖掘与转化智能体 Backend", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# SQLAlchemy 数据库模型
def _now():
    """返回东八区当前时间（无时区信息，值为东八区时间）"""
    return datetime.now(timezone(timedelta(hours=8))).replace(tzinfo=None)


class DBLead(Base):
    __tablename__ = "tbl_leads"
    id = Column(String(64), primary_key=True, index=True)
    title = Column(String(255), nullable=False)
    company = Column(String(255), nullable=False)
    publish_date = Column(Date, nullable=False, default=date.today)
    region = Column(Text, nullable=False)
    industry = Column(String(100))
    category = Column(String(100))
    source_db = Column(String(150))
    requirements = Column(Text)
    notes = Column(Text)
    
    budget_score = Column(Integer, default=50)
    match_score = Column(Integer, default=50)
    stage_score = Column(Integer, default=50)
    qualification_score = Column(Integer, default=50)
    region_score = Column(Integer, default=50)
    total_score = Column(Numeric(5, 2))
    grade = Column(String(1), default="B")
    
    deadline_date = Column(Date, nullable=True)
    budget_amount = Column(Numeric(12, 2), nullable=True)
    is_urgent = Column(Integer, default=0) # 通过整数表示布尔值
    labels = Column(Text, nullable=True)
    crawl_time = Column(DateTime, nullable=True) # 抓取时间
    lead_source = Column(String(20), default="crawler")  # 线索来源: crawler / import / manual
    created_at = Column(DateTime, default=_now)
    
    status = Column(String(30), default="待分级")
    follow_up_person = Column(String(100))
    contact_name = Column(String(100))
    contact_phone = Column(String(30))
    return_reason = Column(Text)
    
    pre_research_report = Column(Text)
    research_date = Column(Date)
    proposal_draft = Column(Text)
    proposal_price = Column(Numeric(12, 2))
    competitor_analysis = Column(Text)
    created_by = Column(String(100))

class DBUser(Base):
    __tablename__ = "tbl_users"
    id = Column(Integer, primary_key=True, autoincrement=True)
    username = Column(String(100), nullable=False, unique=True, index=True)
    password = Column(String(255), nullable=False)
    name = Column(String(100), nullable=False)
    role = Column(String(50), nullable=False)
    created_at = Column(DateTime, default=_now)

class DBUserFavorite(Base):
    __tablename__ = "tbl_user_favorites"
    id = Column(Integer, primary_key=True, autoincrement=True)
    user_id = Column(Integer, nullable=False, index=True)
    lead_id = Column(String(64), nullable=False, index=True)

class DBScoringRule(Base):
    __tablename__ = "tbl_scoring_rules"
    id = Column(Integer, primary_key=True, autoincrement=True)
    budget_weight = Column(Integer, default=20)
    match_weight = Column(Integer, default=30)
    stage_weight = Column(Integer, default=25)
    qualification_weight = Column(Integer, default=15)
    region_weight = Column(Integer, default=10)
    warning_threshold = Column(Integer, default=80)

class DBSearchConfig(Base):
    __tablename__ = "tbl_search_config"
    id = Column(Integer, primary_key=True, autoincrement=True)
    rule_name = Column(String(100), default="默认规则")
    status = Column(String(20), default="active")
    monitor_sources = Column(Text, default="[]")
    keywords = Column(Text, default="[]")
    exclude_keywords = Column(Text, default="[]")
    regions = Column(Text, default="[]")
    industries = Column(Text, default="[]")
    categories = Column(Text, default="[]")
    frequency = Column(String(100), default="每天早上8:00实时")
    budget_min = Column(Integer, default=0)
    budget_max = Column(Integer, default=2000)
    purchase_types = Column(Text, default="[]")
    time_range = Column(String(100), default="不限")

class DBAuditLog(Base):
    __tablename__ = "tbl_audit_logs"
    id = Column(String(64), primary_key=True)
    timestamp = Column(DateTime, nullable=False, default=_now)
    operator = Column(String(100), nullable=False)
    role = Column(String(50), nullable=False)
    action_type = Column(String(100), nullable=False)
    target_id = Column(String(64), nullable=True)
    details = Column(Text)

class DBCompetitor(Base):
    __tablename__ = "tbl_competitors"
    id = Column(String(64), primary_key=True)
    name = Column(String(255), nullable=False)
    strengths = Column(Text)
    weaknesses = Column(Text)
    pricing_range = Column(String(100))
    win_rate = Column(Integer, default=50)
    main_business = Column(Text)            # 主营业务
    tech_route = Column(Text)               # 技术路线
    typical_cases = Column(Text)            # 典型案例（纯文本或 JSON 文本）
    info_source = Column(Text)              # 信息来源（来源链接/出处）
    monitoring_intel = Column(Text)         # JSON: 最新监测动态(新品/中标/技术/报价/案例)
    updated_at = Column(DateTime, default=_now)  # 更新时间
    status = Column(String(20), default="active")           # active=有效 / inactive=失效
    priority = Column(String(10), default="normal")         # key=重点关注 / normal=普通
    last_enriched_at = Column(DateTime, nullable=True)      # 最近一次 AI 检索时间（调度用）

class DBCompetitorUpdate(Base):
    """竞品动态更新记录：每次 AI(DeepSeek) 全网检索时留痕，可追溯。"""
    __tablename__ = "tbl_competitor_updates"
    id = Column(String(64), primary_key=True)
    competitor_id = Column(String(64), index=True, nullable=False)
    changed_fields = Column(Text)        # JSON: 本次变更的基础字段摘要
    intel_payload = Column(Text)         # JSON: 五类监测动态完整内容
    source = Column(Text)                # 信息来源（DeepSeek/检索出处）
    created_at = Column(DateTime, default=_now)

class DBTemplate(Base):
    __tablename__ = "tbl_templates"
    id = Column(String(64), primary_key=True)
    title = Column(String(255), nullable=False)
    category = Column(String(100), nullable=False)
    description = Column(Text)
    content = Column(Text)

class DBResearchReport(Base):
    """预调研档案：每次生成保留为一个历史版本，可追溯。"""
    __tablename__ = "tbl_research_reports"
    id = Column(String(64), primary_key=True)
    lead_id = Column(String(64), index=True, nullable=False)
    version = Column(Integer, default=1)
    created_at = Column(DateTime, default=_now)
    operator = Column(String(100))
    engine = Column(String(20))            # 'rule' | 'llm'
    completeness = Column(Integer, default=0)   # 信息完整度 0-100
    aggregated_info = Column(Text)         # JSON: 结构化调研原始信息集（含来源）
    sources = Column(Text)                 # JSON: [{label, url}]
    missing_items = Column(Text)           # JSON: list[str] 缺失/待确认项
    report_content = Column(Text)          # 渲染后的痛点分析报告全文
    is_sufficient = Column(Integer, default=1)  # 0=公开信息不足

class DBLeadActivity(Base):
    """线索活动/跟进记录表"""
    __tablename__ = "tbl_lead_activities"
    id = Column(String(64), primary_key=True)
    lead_id = Column(String(64), index=True, nullable=False)
    activity_type = Column(String(50), default="comment")  # comment / status_change / note / system
    content = Column(Text)
    operator = Column(String(100))
    created_at = Column(DateTime, default=_now)

class DBCustomerProfile(Base):
    """线索转化生成的客户档案"""
    __tablename__ = "tbl_customer_profiles"
    id = Column(String(64), primary_key=True)
    lead_id = Column(String(64), nullable=False, index=True)
    company = Column(String(255), nullable=False)
    industry = Column(String(100))
    region = Column(String(100))
    contact_name = Column(String(100))
    contact_phone = Column(String(30))
    source_db = Column(String(150))
    budget_amount = Column(Numeric(12, 2))
    requirements = Column(Text)
    converted_by = Column(String(100))
    converted_at = Column(DateTime)
    created_at = Column(DateTime, default=_now)

class DBProjectLedger(Base):
    """线索转化生成的项目台账"""
    __tablename__ = "tbl_project_ledgers"
    id = Column(String(64), primary_key=True)
    lead_id = Column(String(64), nullable=False, index=True)
    customer_profile_id = Column(String(64), nullable=False)
    project_name = Column(String(255), nullable=False)
    project_amount = Column(Numeric(12, 2))
    project_stage = Column(String(50), default="立项")
    region = Column(String(100))
    status = Column(String(20), default="active")
    created_by = Column(String(100))
    created_at = Column(DateTime, default=_now)

class DBSourceHealth(Base):
    """数据源健康监控"""
    __tablename__ = "tbl_source_health"
    id = Column(Integer, primary_key=True, autoincrement=True)
    source = Column(String(255), nullable=False, unique=True, index=True)
    status = Column(String(20), default="healthy")    # healthy / degraded / abnormal
    consecutive_failures = Column(Integer, default=0)
    total_fetches = Column(Integer, default=0)
    total_failures = Column(Integer, default=0)
    structure_change_detected = Column(Integer, default=0)
    last_success_at = Column(DateTime, nullable=True)
    last_failure_at = Column(DateTime, nullable=True)
    last_error_message = Column(Text)
    created_at = Column(DateTime, default=_now)

class DBPushRule(Base):
    """推送规则配置"""
    __tablename__ = "tbl_push_rules"
    id = Column(String(64), primary_key=True)
    rule_name = Column(String(100), default="默认推送规则")
    status = Column(String(20), default="active")
    high_intent_mode = Column(String(50), default="realtime")   # realtime / batch
    medium_intent_schedule_hour = Column(Integer, default=18)
    medium_intent_schedule_minute = Column(Integer, default=0)
    low_intent_silent = Column(Integer, default=1)
    targets = Column(Text, default="[]")   # JSON: [{receiverName, receiverRole, channels}]
    created_at = Column(DateTime, default=_now)

class DBPushRecord(Base):
    """推送记录"""
    __tablename__ = "tbl_push_records"
    id = Column(String(64), primary_key=True)
    title = Column(String(255), nullable=False)
    receiver_name = Column(String(100))
    channels = Column(String(100))        # email / sms / wechat / system
    status = Column(String(20), default="pending")  # pending / retrying / sent / failed
    retry_count = Column(Integer, default=0)
    max_retry = Column(Integer, default=3)
    error_message = Column(Text)
    content = Column(Text)
    created_at = Column(DateTime, default=_now)

# 将新表纳入自动创建
for model in [DBSourceHealth, DBPushRule, DBPushRecord]:
    if not hasattr(model, '__table__') or not hasattr(model.__table__, 'key') or model.__table__.key not in [t.key for t in Base.metadata.sorted_tables]:
        pass  # create_all handles this

# 自动创建物理表结构
Base.metadata.create_all(bind=engine)


def _ensure_competitor_columns() -> None:
    """兼容已存在的旧 tbl_competitors：create_all 不会为既有表补列，
    这里检测并补齐竞品库新增字段，避免破坏既有数据与查询。"""
    new_columns = {
        "main_business": "ADD COLUMN `main_business` TEXT",
        "tech_route": "ADD COLUMN `tech_route` TEXT",
        "typical_cases": "ADD COLUMN `typical_cases` TEXT",
        "info_source": "ADD COLUMN `info_source` TEXT",
        "updated_at": "ADD COLUMN `updated_at` DATETIME NULL",
        "status": "ADD COLUMN `status` VARCHAR(20) DEFAULT 'active'",
        "monitoring_intel": "ADD COLUMN `monitoring_intel` TEXT",
        "priority": "ADD COLUMN `priority` VARCHAR(10) DEFAULT 'normal'",
        "last_enriched_at": "ADD COLUMN `last_enriched_at` DATETIME NULL",
    }
    update_columns = {
        "intel_payload": "ADD COLUMN `intel_payload` TEXT",
    }
    try:
        from sqlalchemy import inspect as _sa_inspect, text as _sa_text
        inspector = _sa_inspect(engine)
        if "tbl_competitors" in inspector.get_table_names():
            existing = {col["name"] for col in inspector.get_columns("tbl_competitors")}
            missing = [ddl for name, ddl in new_columns.items() if name not in existing]
            if missing:
                with engine.begin() as conn:
                    conn.execute(_sa_text(f"ALTER TABLE `tbl_competitors` {', '.join(missing)}"))
                logger.info("tbl_competitors 自动补列完成：%s", [c.split('`')[1] for c in missing])
        if "tbl_competitor_updates" in inspector.get_table_names():
            existing_up = {col["name"] for col in inspector.get_columns("tbl_competitor_updates")}
            missing_up = [ddl for name, ddl in update_columns.items() if name not in existing_up]
            if missing_up:
                with engine.begin() as conn:
                    conn.execute(_sa_text(f"ALTER TABLE `tbl_competitor_updates` {', '.join(missing_up)}"))
                logger.info("tbl_competitor_updates 自动补列完成：%s", [c.split('`')[1] for c in missing_up])
        if "tbl_leads" in inspector.get_table_names():
            existing_leads = {col["name"] for col in inspector.get_columns("tbl_leads")}
            if "created_by" not in existing_leads:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_leads` ADD COLUMN `created_by` VARCHAR(100) NULL"))
                logger.info("tbl_leads 自动补列 created_by 完成")
            if "return_reason" not in existing_leads:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_leads` ADD COLUMN `return_reason` TEXT NULL"))
                logger.info("tbl_leads 自动补列 return_reason 完成")
            if "lead_source" not in existing_leads:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_leads` ADD COLUMN `lead_source` VARCHAR(20) DEFAULT 'crawler' NULL"))
                logger.info("tbl_leads 自动补列 lead_source 完成")
            if "created_at" not in existing_leads:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_leads` ADD COLUMN `created_at` DATETIME NULL"))
                logger.info("tbl_leads 自动补列 created_at 完成")
            region_col = next((c for c in inspector.get_columns("tbl_leads") if c["name"] == "region"), None)
            if region_col and str(region_col.get("type", "")).upper().startswith("VARCHAR"):
                with engine.begin() as conn:
                    conn.execute(_sa_text(
                        "ALTER TABLE `tbl_leads` MODIFY COLUMN `region` TEXT NOT NULL COMMENT '行政区域'"
                    ))
                logger.info("tbl_leads region 字段已扩展为 TEXT")
        if "tbl_customer_profiles" in inspector.get_table_names():
            existing_cust = {col["name"] for col in inspector.get_columns("tbl_customer_profiles")}
            if "converted_by" not in existing_cust:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_customer_profiles` ADD COLUMN `converted_by` VARCHAR(100) NULL"))
                logger.info("tbl_customer_profiles 自动补列 converted_by 完成")
            if "converted_at" not in existing_cust:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_customer_profiles` ADD COLUMN `converted_at` DATETIME NULL"))
                logger.info("tbl_customer_profiles 自动补列 converted_at 完成")
        if "tbl_project_ledgers" in inspector.get_table_names():
            existing_proj = {col["name"] for col in inspector.get_columns("tbl_project_ledgers")}
            if "created_by" not in existing_proj:
                with engine.begin() as conn:
                    conn.execute(_sa_text("ALTER TABLE `tbl_project_ledgers` ADD COLUMN `created_by` VARCHAR(100) NULL"))
                logger.info("tbl_project_ledgers 自动补列 created_by 完成")
    except Exception as exc:
        logger.warning("竞品表或线索表自动补列失败（可忽略，若已手动建表）：%s", exc)


_ensure_competitor_columns()

# 依赖注入会话辅助函数
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# 辅助函数
def _safe_json_list(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except Exception:
        return [s.strip() for s in raw.splitlines() if s.strip()]


def _safe_json_dict(raw: Optional[str]) -> Dict[str, Any]:
    if not raw:
        return {}
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}

def _default_search_config() -> Dict[str, Any]:
    res = {
        key: list(value) if isinstance(value, list) else value
        for key, value in DEFAULT_SEARCH_CONFIG.items()
    }
    res["id"] = 0
    res["ruleName"] = "默认规则"
    res["status"] = "active"
    return res

def _coerce_list(value: Any, default: Optional[List[str]] = None) -> List[str]:
    if value is None:
        return list(default or [])
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return list(default or [])
        try:
            parsed = json.loads(value)
            if isinstance(parsed, list):
                value = parsed
            else:
                value = value.split(",")
        except json.JSONDecodeError:
            value = value.split(",")
    if isinstance(value, (list, tuple, set)):
        return [str(item).strip() for item in value if str(item).strip()]
    return list(default or [])

def _coerce_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default

def normalize_search_config(raw: Optional[Dict[str, Any]]) -> Dict[str, Any]:
    config = _default_search_config()
    if isinstance(raw, dict):
        config.update(raw)

    for field in ("monitorSources", "keywords", "excludeKeywords", "regions", "industries", "categories", "purchaseTypes"):
        config[field] = _coerce_list(config.get(field), DEFAULT_SEARCH_CONFIG[field])

    # 严格控制监控源：仅限已映射的源名称或显式 URL。
    config["monitorSources"] = [
        source for source in config["monitorSources"]
        if source in SOURCE_URL_MAP or _is_url(source)
    ]

    config["budgetMin"] = _coerce_int(config.get("budgetMin"), DEFAULT_SEARCH_CONFIG["budgetMin"])
    config["budgetMax"] = _coerce_int(config.get("budgetMax"), DEFAULT_SEARCH_CONFIG["budgetMax"])
    if config["budgetMax"] < config["budgetMin"]:
        config["budgetMin"], config["budgetMax"] = config["budgetMax"], config["budgetMin"]

    config["frequency"] = str(config.get("frequency") or DEFAULT_SEARCH_CONFIG["frequency"])
    config["timeRange"] = str(
        config.get("timeRange") or config.get("time_range") or DEFAULT_SEARCH_CONFIG["timeRange"]
    ).strip() or DEFAULT_SEARCH_CONFIG["timeRange"]
    return config

def search_config_to_api(config: Optional[DBSearchConfig]) -> Dict[str, Any]:
    if not config:
        return _default_search_config()
    
    base_config = normalize_search_config({
        "monitorSources": config.monitor_sources,
        "keywords": config.keywords,
        "excludeKeywords": getattr(config, "exclude_keywords", None),
        "regions": config.regions,
        "industries": config.industries,
        "categories": config.categories,
        "frequency": config.frequency,
        "budgetMin": getattr(config, "budget_min", None),
        "budgetMax": getattr(config, "budget_max", None),
        "purchaseTypes": getattr(config, "purchase_types", None),
        "timeRange": getattr(config, "time_range", DEFAULT_SEARCH_CONFIG["timeRange"]),
    })
    
    base_config["id"] = config.id
    base_config["ruleName"] = getattr(config, "rule_name", "默认规则")
    base_config["status"] = getattr(config, "status", "active")
    
    return base_config

def get_or_create_scoring_rule(db: Session) -> DBScoringRule:
    rules = db.query(DBScoringRule).first()
    if rules:
        return rules
    rules = DBScoringRule(
        budget_weight=20,
        match_weight=30,
        stage_weight=25,
        qualification_weight=15,
        region_weight=10,
        warning_threshold=80,
    )
    db.add(rules)
    db.commit()
    db.refresh(rules)
    logger.info("Initialized default scoring rule")
    return rules

def calculate_score(
    rules: DBScoringRule,
    budget_score: int,
    match_score: int,
    stage_score: int,
    qualification_score: int,
    region_score: int,
) -> Tuple[float, str]:
    total = (
        budget_score * (rules.budget_weight / 100.0) +
        match_score * (rules.match_weight / 100.0) +
        stage_score * (rules.stage_weight / 100.0) +
        qualification_score * (rules.qualification_weight / 100.0) +
        region_score * (rules.region_weight / 100.0)
    )
    total = round(float(total), 2)
    grade = "A" if total >= rules.warning_threshold else ("B" if total >= 50 else "C")
    return total, grade

def _date_str(d: Optional[date]) -> Optional[str]:
    return d.isoformat() if d else None


def _dt_str(dt_val: Optional[datetime]) -> Optional[str]:
    if dt_val is None:
        return None
    # 所有存储的时间均为东八区时间（无时区信息），显式添加 +08:00 偏移
    if dt_val.tzinfo is None:
        return dt_val.isoformat() + "+08:00"
    return dt_val.isoformat()

def lead_to_api(lead: DBLead, favorited_ids: set = None) -> dict:
    return {
        "id": lead.id,
        "title": lead.title,
        "company": lead.company,
        "publishDate": _date_str(lead.publish_date) or "",
        "region": _sanitize_region(lead.region) or lead.region or "待核验地区",
        "industry": lead.industry or "",
        "category": lead.category or "",
        "sourceDb": lead.source_db or "",
        "requirements": lead.requirements or "",
        "notes": lead.notes,
        "budgetScore": int(lead.budget_score or 0),
        "matchScore": int(lead.match_score or 0),
        "stageScore": int(lead.stage_score or 0),
        "qualificationScore": int(lead.qualification_score or 0),
        "regionScore": int(lead.region_score or 0),
        "totalScore": float(lead.total_score or 0),
        "grade": lead.grade or "B",
        "deadlineDate": _date_str(lead.deadline_date) or "",
        "budgetAmount": float(lead.budget_amount) if lead.budget_amount is not None else None,
        "isUrgent": bool(lead.is_urgent),
        "labels": lead.labels or "",
        "isFavorite": favorited_ids is not None and lead.id in favorited_ids,
        "crawlTime": _dt_str(lead.crawl_time) if lead.crawl_time else "",
        "source": lead.lead_source or "crawler",
        "createdAt": _dt_str(lead.created_at) if lead.created_at else "",
        "status": lead.status or "待分级",
        "followUpPerson": lead.follow_up_person,
        "contactName": lead.contact_name,
        "contactPhone": lead.contact_phone,
        "preResearchReport": lead.pre_research_report,
        "researchDate": _date_str(lead.research_date),
        "proposalDraft": lead.proposal_draft,
        "proposalPrice": float(lead.proposal_price) if lead.proposal_price is not None else None,
        "competitorAnalysis": lead.competitor_analysis,
        "createdBy": lead.created_by or "",
        "returnReason": lead.return_reason or "",
    }

def rule_to_api(rule: DBScoringRule) -> dict:
    return {
        "budgetWeight": int(rule.budget_weight),
        "matchWeight": int(rule.match_weight),
        "stageWeight": int(rule.stage_weight),
        "qualificationWeight": int(rule.qualification_weight),
        "regionWeight": int(rule.region_weight),
        "warningThreshold": int(rule.warning_threshold),
    }

def audit_to_api(log: DBAuditLog) -> dict:
    return {
        "id": log.id,
        "timestamp": _dt_str(log.timestamp) or "",
        "operator": log.operator,
        "role": log.role,
        "actionType": log.action_type,
        "targetId": log.target_id,
        "details": log.details or "",
    }

def competitor_to_api(comp: DBCompetitor) -> dict:
    return {
        "id": comp.id,
        "name": comp.name,
        "strengths": _safe_json_list(comp.strengths),
        "weaknesses": _safe_json_list(comp.weaknesses),
        "pricingRange": comp.pricing_range or "",
        "winRate": int(comp.win_rate or 0),
        "pastProjects": [],
        "mainBusiness": getattr(comp, "main_business", None) or "",
        "techRoute": getattr(comp, "tech_route", None) or "",
        "typicalCases": getattr(comp, "typical_cases", None) or "",
        "infoSource": getattr(comp, "info_source", None) or "",
        "monitoringIntel": _safe_json_dict(getattr(comp, "monitoring_intel", None)),
        "updatedAt": _dt_str(getattr(comp, "updated_at", None)),
        "status": getattr(comp, "status", None) or "active",
        "priority": getattr(comp, "priority", None) or "normal",
        "lastEnrichedAt": _dt_str(getattr(comp, "last_enriched_at", None)),
    }

def competitor_update_to_api(rec: DBCompetitorUpdate) -> dict:
    try:
        changed = json.loads(rec.changed_fields) if rec.changed_fields else {}
    except Exception:
        changed = {}
    return {
        "id": rec.id,
        "competitorId": rec.competitor_id,
        "changedFields": changed,
        "intelPayload": _safe_json_dict(getattr(rec, "intel_payload", None)),
        "source": rec.source or "",
        "createdAt": _dt_str(rec.created_at),
    }

def template_to_api(tpl: DBTemplate) -> dict:
    return {
        "id": tpl.id,
        "title": tpl.title,
        "category": tpl.category,
        "description": tpl.description or "",
        "content": tpl.content or "",
    }

# 数据模型（Schema）
class LeadCreate(BaseModel):
    title: str
    company: str
    region: str
    publishDate: Optional[str] = None
    industry: Optional[str] = None
    category: Optional[str] = None
    sourceDb: Optional[str] = None
    requirements: Optional[str] = None
    notes: Optional[str] = None
    budgetScore: int = 50
    matchScore: int = 50
    stageScore: int = 50
    qualificationScore: int = 50
    regionScore: int = 50
    deadlineDate: Optional[str] = None
    budgetAmount: Optional[float] = None
    isUrgent: bool = False
    labels: Optional[str] = None
    contactName: Optional[str] = None
    contactPhone: Optional[str] = None
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None

class LeadUpdate(BaseModel):
    title: Optional[str] = None
    company: Optional[str] = None
    publishDate: Optional[str] = None
    region: Optional[str] = None
    industry: Optional[str] = None
    category: Optional[str] = None
    sourceDb: Optional[str] = None
    requirements: Optional[str] = None
    notes: Optional[str] = None
    budgetScore: Optional[int] = None
    matchScore: Optional[int] = None
    stageScore: Optional[int] = None
    qualificationScore: Optional[int] = None
    regionScore: Optional[int] = None
    deadlineDate: Optional[str] = None
    budgetAmount: Optional[float] = None
    isUrgent: Optional[bool] = None
    labels: Optional[str] = None
    status: Optional[str] = None
    followUpPerson: Optional[str] = None
    contactName: Optional[str] = None
    contactPhone: Optional[str] = None
    returnReason: Optional[str] = None
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None

class RuleUpdate(BaseModel):
    budgetWeight: int
    matchWeight: int
    stageWeight: int
    qualificationWeight: int
    regionWeight: int
    warningThreshold: int
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None

class SearchConfigResponse(BaseModel):
    id: int
    ruleName: str
    status: str
    monitorSources: List[str]
    keywords: List[str]
    excludeKeywords: List[str]
    regions: List[str]
    industries: List[str]
    categories: List[str]
    frequency: str
    budgetMin: int
    budgetMax: int
    purchaseTypes: List[str]
    timeRange: str

class SearchConfigUpdate(BaseModel):
    ruleName: str
    status: str
    monitorSources: List[str]
    keywords: List[str]
    excludeKeywords: List[str]
    regions: List[str]
    industries: List[str]
    categories: List[str]
    frequency: str
    budgetMin: int
    budgetMax: int
    purchaseTypes: List[str]
    timeRange: str = "不限"
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None

class CompetitorCreate(BaseModel):
    name: str
    strengths: List[str] = Field(default_factory=list)
    weaknesses: List[str] = Field(default_factory=list)
    pricingRange: Optional[str] = None
    winRate: int = 50
    mainBusiness: Optional[str] = None
    techRoute: Optional[str] = None
    typicalCases: Optional[str] = None
    infoSource: Optional[str] = None
    priority: Optional[str] = "normal"
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None


class CompetitorUpdate(BaseModel):
    name: Optional[str] = None
    strengths: Optional[List[str]] = None
    weaknesses: Optional[List[str]] = None
    pricingRange: Optional[str] = None
    winRate: Optional[int] = None
    mainBusiness: Optional[str] = None
    techRoute: Optional[str] = None
    typicalCases: Optional[str] = None
    infoSource: Optional[str] = None
    status: Optional[str] = None
    priority: Optional[str] = None
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None


class CompetitorStatusUpdate(BaseModel):
    status: str
    operatorName: Optional[str] = None
    operatorRole: Optional[str] = None

# 用户管理数据模型
class UserLogin(BaseModel):
    username: str
    password: str

class UserRegister(BaseModel):
    username: str
    password: str
    name: str
    role: str

class UserCreate(BaseModel):
    username: str
    password: str
    name: str
    role: str

class UserRoleUpdate(BaseModel):
    role: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    username: Optional[str] = None
    password: Optional[str] = None

# 密码哈希辅助函数（SHA-256）
import hashlib
def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode('utf-8')).hexdigest()

# 通过请求头提取当前用户的依赖
from fastapi import Header
from urllib.parse import unquote

def get_current_user(
    x_user_id: Optional[str] = Header(None),
    x_user_username: Optional[str] = Header(None),
    x_user_name: Optional[str] = Header(None),
    x_user_role: Optional[str] = Header(None)
):
    if not x_user_id or not x_user_username or not x_user_name or not x_user_role:
        raise HTTPException(status_code=401, detail="请先登录系统")
    try:
        decoded_username = unquote(x_user_username)
        decoded_name = unquote(x_user_name)
        decoded_role = unquote(x_user_role)
    except Exception:
        decoded_username = x_user_username
        decoded_name = x_user_name
        decoded_role = x_user_role
    return {
        "id": x_user_id,
        "username": decoded_username,
        "name": decoded_name,
        "role": decoded_role
    }

# 如不存在则初始化默认用户
def seed_default_users(db: Session):
    admin_exists = db.query(DBUser).filter(DBUser.username == "admin").first()
    if not admin_exists:
        users_to_seed = [
            DBUser(username="admin", password=hash_password("admin123"), name="超级管理员陈工", role="超级管理员"),
            DBUser(username="manager", password=hash_password("manager123"), name="市场组长刘经理", role="商务负责人"),
            DBUser(username="sales", password=hash_password("sales123"), name="前端突破专员王小二", role="商务专员")
        ]
        db.add_all(users_to_seed)
        db.commit()

# 模块加载时执行用户初始化
_seed_db = SessionLocal()
try:
    seed_default_users(_seed_db)
finally:
    _seed_db.close()

def user_to_api(user: DBUser) -> dict:
    return {
        "id": user.id,
        "username": user.username,
        "name": user.name,
        "role": user.role,
        "createdAt": _dt_str(user.created_at)
    }

# 审计日志辅助函数
def add_audit_log(db: Session, operator: str, role: str, action_type: str, details: str, target_id: Optional[str] = None):
    new_log = DBAuditLog(
        id=str(uuid.uuid4()),
        timestamp=_now(),
        operator=operator,
        role=role,
        action_type=action_type,
        target_id=target_id,
        details=details
    )
    db.add(new_log)
    db.commit()

# --- 认证接口 ---
@app.post("/api/auth/login")
def login(payload: UserLogin, db: Session = Depends(get_db)):
    if not payload.username or not payload.password:
        raise HTTPException(status_code=400, detail="用户名和密码不能为空")
    user = db.query(DBUser).filter(DBUser.username == payload.username).first()
    if not user:
        raise HTTPException(status_code=401, detail="用户名不存在")
    if user.password != hash_password(payload.password):
        raise HTTPException(status_code=401, detail="密码错误")
    return user_to_api(user)

@app.post("/api/auth/register", status_code=201)
def register(payload: UserRegister, db: Session = Depends(get_db)):
    if not payload.username or not payload.password or not payload.name:
        raise HTTPException(status_code=400, detail="必填项不能为空")
    
    # 检查重复
    existing = db.query(DBUser).filter(DBUser.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=409, detail="用户名已存在")
        
    new_user = DBUser(
        username=payload.username,
        password=hash_password(payload.password),
        name=payload.name,
        role="商务专员", # 注册默认角色为“商务专员”，后续由超级管理员进行权限分配
        created_at=_now()
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    add_audit_log(
        db,
        new_user.name,
        new_user.role,
        "用户注册",
        f"新用户注册成功，用户名: {new_user.username}",
        new_user.id
    )
    return user_to_api(new_user)

# --- 用户管理接口（仅超级管理员） ---
@app.get("/api/users")
def get_users(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权查看用户列表")
    users_list = db.query(DBUser).all()
    return [user_to_api(u) for u in users_list]

@app.post("/api/users", status_code=201)
def create_system_user(payload: UserCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "超级管理员":
        raise HTTPException(status_code=403, detail="仅超级管理员有权创建用户")
    if not payload.username or not payload.password or not payload.name or not payload.role:
        raise HTTPException(status_code=400, detail="必填项不能为空")
        
    # 检查重复
    existing = db.query(DBUser).filter(DBUser.username == payload.username).first()
    if existing:
        raise HTTPException(status_code=409, detail="用户名已存在")
        
    new_user = DBUser(
        username=payload.username,
        password=hash_password(payload.password),
        name=payload.name,
        role=payload.role,
        created_at=_now()
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "创建用户",
        f"管理员创建了新用户: {new_user.name} ({new_user.role})",
        new_user.id
    )
    return user_to_api(new_user)

@app.put("/api/users/{user_id}/role")
def update_user_role(user_id: int, payload: UserRoleUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "超级管理员":
        raise HTTPException(status_code=403, detail="仅超级管理员有权修改用户角色")
    if not payload.role:
        raise HTTPException(status_code=400, detail="角色不能为空")
        
    target_user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="用户不存在")
        
    old_role = target_user.role
    target_user.role = payload.role
    db.commit()
    db.refresh(target_user)
    
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "修改用户角色",
        f"将用户 {target_user.name} 的角色由 「{old_role}」 修改为 「{payload.role}」",
        target_user.id
    )
    return user_to_api(target_user)

@app.delete("/api/users/{user_id}")
def delete_system_user(user_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "超级管理员":
        raise HTTPException(status_code=403, detail="仅超级管理员有权删除用户")
    
    target_user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="用户不存在")
        
    if target_user.username == "admin":
        raise HTTPException(status_code=400, detail="不能删除初始超级管理员账号")

    # 校验：该用户是否仍有「跟进中」的线索
    following_count = db.query(DBLead).filter(
        DBLead.status == "跟进中",
        (DBLead.follow_up_person == target_user.name) | (DBLead.follow_up_person == target_user.username)
    ).count()
    if following_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"该用户仍有 {following_count} 条跟进中的线索，请先转移跟进人或完结线索后再删除"
        )

    user_name = target_user.name
    user_role = target_user.role
    db.delete(target_user)
    db.commit()
    
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "删除用户",
        f"删除了用户: {user_name} ({user_role})",
        user_id
    )
    return {"success": True}

@app.put("/api/users/{user_id}")
def update_system_user(user_id: int, payload: UserUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "超级管理员":
        raise HTTPException(status_code=403, detail="仅超级管理员有权修改用户信息")
    if user_id == "user-admin" and current_user.get("id") != "user-admin":
        raise HTTPException(status_code=403, detail="不能修改初始超级管理员账号")

    target_user = db.query(DBUser).filter(DBUser.id == user_id).first()
    if not target_user:
        raise HTTPException(status_code=404, detail="用户不存在")

    changed = []
    if payload.name is not None and payload.name != target_user.name:
        changed.append(f"姓名: {target_user.name} → {payload.name}")
        target_user.name = payload.name
    if payload.username is not None and payload.username != target_user.username:
        existing = db.query(DBUser).filter(DBUser.username == payload.username, DBUser.id != user_id).first()
        if existing:
            raise HTTPException(status_code=409, detail=f"用户名「{payload.username}」已被占用")
        changed.append(f"用户名: {target_user.username} → {payload.username}")
        target_user.username = payload.username
    if payload.password is not None and payload.password.strip():
        new_hash = hash_password(payload.password)
        if new_hash != target_user.password:
            changed.append("密码已修改")
            target_user.password = new_hash

    if not changed:
        raise HTTPException(status_code=400, detail="未检测到任何修改")

    db.commit()
    db.refresh(target_user)

    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "修改用户信息",
        f"修改用户 {target_user.name}：{'；'.join(changed)}",
        target_user.id
    )
    return user_to_api(target_user)

# API 接口
@app.get("/api/leads")
def read_leads(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role in ("超级管理员", "商务负责人"):
        leads = db.query(DBLead).order_by(DBLead.publish_date.desc()).all()
    elif role == "商务专员":
        # 商务专员只能看到自己负责的或者自己创建的线索
        leads = db.query(DBLead).filter(
            (DBLead.follow_up_person == current_user["name"]) | 
            (DBLead.created_by == current_user["username"])
        ).order_by(DBLead.publish_date.desc()).all()
    elif role == "查看角色":
        leads = db.query(DBLead).order_by(DBLead.publish_date.desc()).all()
    else:
        leads = []
    
    # 查询当前用户收藏的线索 ID 集合
    uid = int(current_user["id"])
    fav_rows = db.query(DBUserFavorite.lead_id).filter(DBUserFavorite.user_id == uid).all()
    favorited_ids = {row[0] for row in fav_rows}
    
    result = []
    for l in leads:
        api_lead = lead_to_api(l, favorited_ids)
        if role == "查看角色":
            # 联系人姓名：隐藏或置空（如"无权查看"）
            api_lead["contactName"] = "无权查看"
            # 联系人电话：隐藏或置空（如"无权查看"）
            api_lead["contactPhone"] = "无权查看"
            # 方案报价/预算：隐藏或置空（如 0 或 null）
            api_lead["proposalPrice"] = None
        result.append(api_lead)
        
    return result

@app.post("/api/leads")
def create_lead(lead: LeadCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role == "查看角色":
        raise HTTPException(status_code=403, detail="查看角色无权创建线索")

    rules = get_or_create_scoring_rule(db)
    total, grade = calculate_score(
        rules,
        lead.budgetScore,
        lead.matchScore,
        lead.stageScore,
        lead.qualificationScore,
        lead.regionScore,
    )

    publish_date = date.today()
    if lead.publishDate:
        try:
            publish_date = date.fromisoformat(lead.publishDate)
        except Exception:
            publish_date = date.today()
    
    deadline_date = None
    if lead.deadlineDate:
        try:
            deadline_date = date.fromisoformat(lead.deadlineDate)
        except Exception:
            pass

    db_lead = DBLead(
        id=str(uuid.uuid4()),
        title=lead.title,
        company=lead.company,
        publish_date=publish_date,
        region=lead.region,
        industry=lead.industry,
        category=lead.category,
        source_db=lead.sourceDb or "API/Crawler",
        requirements=lead.requirements,
        notes=lead.notes,
        budget_score=lead.budgetScore,
        match_score=lead.matchScore,
        stage_score=lead.stageScore,
        qualification_score=lead.qualificationScore,
        region_score=lead.regionScore,
        total_score=total,
        grade=grade,
        deadline_date=deadline_date,
        budget_amount=lead.budgetAmount,
        is_urgent=1 if lead.isUrgent else 0,
        labels=lead.labels,
        contact_name=lead.contactName,
        contact_phone=lead.contactPhone,
        status="待分级",
        created_by=current_user["username"],
        lead_source="manual",
        created_at=_now()
    )
    # 打分/评级完成后自动流转
    db_lead.status = "待分配"
    db.add(db_lead)
    db.commit()
    db.refresh(db_lead)
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "新增线索",
        f"新增线索「{db_lead.title}」",
        db_lead.id,
    )
    return lead_to_api(db_lead)

@app.put("/api/leads/{lead_id}")
def update_lead(lead_id: str, patch: LeadUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role == "查看角色":
        raise HTTPException(status_code=403, detail="查看角色无权修改线索")

    db_lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not db_lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    # 数据隔离检查：商务专员只能操作自己负责或创建的线索
    if role == "商务专员" and db_lead.follow_up_person != current_user["name"] and db_lead.created_by != current_user["username"]:
        raise HTTPException(status_code=403, detail="您只能修改您负责或创建的线索")

    # 商务专员无权修改招投标联系人 / 电话
    if role == "商务专员" and (patch.contactName is not None or patch.contactPhone is not None):
        raise HTTPException(status_code=403, detail="商务专员无权修改联系人信息")

    old_status = db_lead.status or "待分级"
    old_follow_up_person = db_lead.follow_up_person or ""
    proposed_status = patch.status if patch.status is not None else old_status
    proposed_follow_up_person = patch.followUpPerson

    valid_states = ["待分级", "待分配", "跟进中", "已转化", "已无效", "已流失"]
    if patch.status is not None and patch.status not in valid_states:
        raise HTTPException(status_code=400, detail=f"无效的状态值: {patch.status}")

    # 禁止通过 PUT 接口将状态改为「已转化」，必须使用专用的 /convert 接口
    if patch.status == "已转化":
        raise HTTPException(status_code=400, detail="禁止通过状态变更改为「已转化」，请使用「转为客户」按钮操作")

    # 自动流转：分配跟进人后从「待分配」转为「跟进中」
    if old_status == "待分配" and proposed_follow_up_person and proposed_follow_up_person.strip():
        proposed_status = "跟进中"

    return_reason_log = ""

    if old_status != proposed_status:
        # 1. 「已转化」完全不可逆
        if old_status == "已转化":
            raise HTTPException(status_code=400, detail="「已转化」状态的线索已锁定，无法再次变更状态")

        # 2. 超级管理员 / 商务负责人可将「已无效」或「已流失」退回至「跟进中」
        if old_status in ("已无效", "已流失"):
            if proposed_status == "跟进中":
                if role not in ("超级管理员", "商务负责人"):
                    raise HTTPException(status_code=403, detail="只有超级管理员或商务负责人有权将「已无效」或「已流失」状态的线索退回到「跟进中」")
                if not patch.returnReason or not patch.returnReason.strip():
                    raise HTTPException(status_code=400, detail="退回操作必须填写退回原因")
                return_reason_log = f"，退回原因：{patch.returnReason.strip()}"
            else:
                raise HTTPException(status_code=400, detail="「已无效」或「已流失」状态仅能由超级管理员退回至「跟进中」状态，不能变更至其他状态")

        # 3. 默认不允许回退
        if old_status == "跟进中" and proposed_status in ("待分配", "待分级"):
            raise HTTPException(status_code=400, detail="跟进中的线索无法回退至「待分配」或「待分级」状态")
        if old_status == "待分配" and proposed_status == "待分级":
            raise HTTPException(status_code=400, detail="待分配的线索无法回退至「待分级」状态")

        # 4. 验证正向流转路径
        is_finalizing = proposed_status in ("已无效", "已流失")
        can_finalize = role in ("超级管理员", "商务负责人")
        if old_status == "待分级" and proposed_status != "待分配":
            if not (can_finalize and is_finalizing):
                raise HTTPException(status_code=400, detail="「待分级」状态的线索在完成打分/评级前，无法直接流转到其他状态")

        if old_status == "待分配":
            if proposed_status == "跟进中":
                check_follow_up = proposed_follow_up_person or db_lead.follow_up_person
                if not check_follow_up or not check_follow_up.strip():
                    raise HTTPException(status_code=400, detail="流转至「跟进中」状态前，必须先分配跟进人")
            elif not (can_finalize and is_finalizing):
                raise HTTPException(status_code=400, detail="「待分配」状态的线索只能在分配跟进人后流转至「跟进中」")

        if old_status == "跟进中" and proposed_status not in ("已转化", "已无效", "已流失"):
            raise HTTPException(status_code=400, detail="「跟进中」状态只能流转至「已转化」、「已无效」或「已流失」")

    if patch.title is not None:
        db_lead.title = patch.title
    if patch.company is not None:
        db_lead.company = patch.company
    if patch.publishDate is not None:
        try:
            db_lead.publish_date = date.fromisoformat(patch.publishDate)
        except Exception:
            pass
    if patch.region is not None:
        db_lead.region = patch.region
    if patch.industry is not None:
        db_lead.industry = patch.industry
    if patch.category is not None:
        db_lead.category = patch.category
    if patch.sourceDb is not None:
        db_lead.source_db = patch.sourceDb
    if patch.requirements is not None:
        db_lead.requirements = patch.requirements
    if patch.notes is not None:
        db_lead.notes = patch.notes
    score_changed = False
    if patch.budgetScore is not None:
        db_lead.budget_score = patch.budgetScore
        score_changed = True
    if patch.matchScore is not None:
        db_lead.match_score = patch.matchScore
        score_changed = True
    if patch.stageScore is not None:
        db_lead.stage_score = patch.stageScore
        score_changed = True
    if patch.qualificationScore is not None:
        db_lead.qualification_score = patch.qualificationScore
        score_changed = True
    if patch.regionScore is not None:
        db_lead.region_score = patch.regionScore
        score_changed = True
    if patch.deadlineDate is not None:
        try:
            db_lead.deadline_date = date.fromisoformat(patch.deadlineDate)
        except Exception:
            pass
    if patch.budgetAmount is not None:
        db_lead.budget_amount = patch.budgetAmount
    if patch.isUrgent is not None:
        db_lead.is_urgent = 1 if patch.isUrgent else 0
    if patch.labels is not None:
        db_lead.labels = patch.labels
    db_lead.status = proposed_status
    if return_reason_log:
        db_lead.return_reason = patch.returnReason.strip()
    if patch.followUpPerson is not None:
        db_lead.follow_up_person = patch.followUpPerson
    if patch.contactName is not None:
        db_lead.contact_name = patch.contactName
    if patch.contactPhone is not None:
        db_lead.contact_phone = patch.contactPhone

    if score_changed:
        rules = get_or_create_scoring_rule(db)
        total, grade = calculate_score(
            rules,
            int(db_lead.budget_score or 0),
            int(db_lead.match_score or 0),
            int(db_lead.stage_score or 0),
            int(db_lead.qualification_score or 0),
            int(db_lead.region_score or 0),
        )
        db_lead.total_score = total
        db_lead.grade = grade
        if not patch.status and (db_lead.status or "待分级") == "待分级":
            db_lead.status = "待分配"

    db.commit()
    db.refresh(db_lead)

    if patch.operatorName or patch.operatorRole:
        log_action = "更新线索"
        details_log = f"更新线索「{db_lead.title}」"
        if old_status != proposed_status:
            if return_reason_log:
                log_action = "线索退回"
                details_log = f"将跟进状态由 「{old_status}」 退回为 「{proposed_status}」{return_reason_log}"
            else:
                log_action = "更新跟进状态"
                details_log = f"将跟进状态由 「{old_status}」 变更为 「{proposed_status}」"
                if db_lead.follow_up_person:
                    details_log += f"，负责人：{db_lead.follow_up_person}"
        add_audit_log(
            db,
            patch.operatorName or "操作员",
            patch.operatorRole or "系统管理员",
            log_action,
            details_log,
            db_lead.id,
        )

        # 自动为状态/分配变更创建活动记录
        activity_content = ""
        activity_type = "system"
        if old_status != proposed_status:
            if return_reason_log:
                activity_content = f"🔄 退回跟进状态: 「{old_status}」→「{proposed_status}」，原因：{patch.returnReason.strip()}"
            else:
                activity_content = f"📌 状态变更: 「{old_status}」→「{proposed_status}」"
        if patch.followUpPerson is not None and patch.followUpPerson != old_follow_up_person:
            if activity_content:
                activity_content += f"\n👤 跟进人指派: {patch.followUpPerson}"
            else:
                activity_content = f"👤 跟进人指派: {patch.followUpPerson}"
        if activity_content:
            db_activity = DBLeadActivity(
                id=str(uuid.uuid4()),
                lead_id=db_lead.id,
                activity_type=activity_type,
                content=activity_content,
                operator=patch.operatorName or "系统",
                created_at=_now(),
            )
            db.add(db_activity)
            db.commit()

    uid = int(current_user["id"])
    fav = db.query(DBUserFavorite.lead_id).filter(
        DBUserFavorite.user_id == uid,
        DBUserFavorite.lead_id == db_lead.id
    ).first()
    return lead_to_api(db_lead, {db_lead.id} if fav else set())

@app.delete("/api/leads/{lead_id}")
def delete_lead(lead_id: str, db: Session = Depends(get_db)):
    db_lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not db_lead:
        raise HTTPException(status_code=404, detail="Lead not found")

    title = db_lead.title
    db.delete(db_lead)
    db.commit()

    add_audit_log(db, "操作员", "系统管理员", "删除线索", f"删除线索「{title}」", lead_id)
    return {"success": True}

# ============ 线索活动（跟进记录） ============
def activity_to_api(a: DBLeadActivity) -> dict:
    return {
        "id": a.id,
        "leadId": a.lead_id,
        "activityType": a.activity_type,
        "content": a.content,
        "operator": a.operator,
        "createdAt": _dt_str(a.created_at) if a.created_at else None,
    }

@app.get("/api/leads/{lead_id}/activities")
def list_lead_activities(lead_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """获取指定线索的所有跟进记录"""
    rows = (
        db.query(DBLeadActivity)
        .filter(DBLeadActivity.lead_id == lead_id)
        .order_by(DBLeadActivity.created_at.asc())
        .all()
    )
    return [activity_to_api(r) for r in rows]

class ActivityCreate(BaseModel):
    activityType: str = "comment"  # comment / system
    content: str
    operatorName: Optional[str] = None

@app.post("/api/leads/{lead_id}/activities", status_code=201)
def create_lead_activity(lead_id: str, payload: ActivityCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """添加一条跟进记录"""
    # 验证线索存在
    lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")

    operator = payload.operatorName or current_user.get("name", "操作员")
    new_activity = DBLeadActivity(
        id=str(uuid.uuid4()),
        lead_id=lead_id,
        activity_type=payload.activityType or "comment",
        content=payload.content,
        operator=operator,
        created_at=_now(),
    )
    db.add(new_activity)
    db.commit()
    db.refresh(new_activity)

    # 操作日志：图片地址以「上传了X张图片」展示
    img_count = len(re.findall(r'!\[[^\]]*\]\(\s*[^)]*\s*\)', payload.content))
    content_display = re.sub(r'!\[[^\]]*\]\(\s*[^)]*\s*\)', '', payload.content).strip()
    if img_count > 0:
        prefix = f"上传了{img_count}张图片"
        content_display = f"{prefix}，{content_display}" if content_display else prefix

    add_audit_log(
        db,
        operator,
        current_user.get("role", "商务专员"),
        "线索跟进",
        f"添加跟进记录: {content_display[:50]}{'...' if len(content_display) > 50 else ''}",
        lead_id,
    )
    return activity_to_api(new_activity)


# 允许上传的图片格式
_ALLOWED_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"}

@app.post("/api/upload/image")
async def upload_image(file: UploadFile = File(...)):
    """上传图片到 MinIO，返回可访问的 URL"""
    client = _get_minio_client()
    if not client:
        raise HTTPException(status_code=503, detail="MinIO 未配置或不可用")

    # 校验文件类型
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in _ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"不支持的图片格式: {ext}，仅支持 {', '.join(sorted(_ALLOWED_IMAGE_EXTENSIONS))}")

    # 生成唯一文件名
    object_name = f"lead_images/{uuid.uuid4().hex}{ext}"

    try:
        content = await file.read()
        result = client.put_object(
            MINIO_BUCKET_NAME,
            object_name,
            io.BytesIO(content),
            length=len(content),
            content_type=file.content_type or "image/jpeg",
        )
        # 构造可访问 URL
        url = f"{MINIO_ENDPOINT}/{MINIO_BUCKET_NAME}/{object_name}"
        return {"url": url, "objectName": object_name}
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"上传失败: {str(exc)}")


# --- 收藏/取消收藏线索（所有角色可用） ---
@app.post("/api/leads/{lead_id}/toggle-favorite")
def toggle_lead_favorite(lead_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "查看角色":
        raise HTTPException(status_code=403, detail="查看角色无权操作")
    db_lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not db_lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    uid = int(current_user["id"])
    existing = db.query(DBUserFavorite).filter(
        DBUserFavorite.user_id == uid,
        DBUserFavorite.lead_id == lead_id
    ).first()
    if existing:
        db.delete(existing)
        is_fav = False
    else:
        db.add(DBUserFavorite(user_id=uid, lead_id=lead_id))
        is_fav = True
    db.commit()
    return lead_to_api(db_lead, {lead_id} if is_fav else set())


# --- 线索转化（生成客户 + 项目台账） ---
@app.post("/api/leads/{lead_id}/convert", status_code=201)
def convert_lead_to_customer(lead_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    """将「跟进中」状态的线索转化为客户档案 + 项目台账"""
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人", "商务专员"):
        raise HTTPException(status_code=403, detail="仅超级管理员、商务负责人或商务专员有权执行线索转化")

    db_lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not db_lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    if db_lead.status != "跟进中":
        raise HTTPException(status_code=400, detail=f"只有「跟进中」状态的线索才能转化，当前状态: {db_lead.status}")

    operator = current_user.get("name", "操作员")

    # 1) 创建客户档案
    customer = DBCustomerProfile(
        id=str(uuid.uuid4()),
        lead_id=db_lead.id,
        company=db_lead.company,
        industry=db_lead.industry,
        region=db_lead.region,
        contact_name=db_lead.contact_name,
        contact_phone=db_lead.contact_phone,
        source_db=db_lead.source_db,
        budget_amount=db_lead.budget_amount,
        requirements=db_lead.requirements,
        converted_by=operator,
        converted_at=_now(),
        created_at=_now(),
    )
    db.add(customer)
    db.flush()

    # 2) 创建项目台账
    project = DBProjectLedger(
        id=str(uuid.uuid4()),
        lead_id=db_lead.id,
        customer_profile_id=customer.id,
        project_name=db_lead.title,
        project_amount=db_lead.budget_amount,
        project_stage="立项",
        region=db_lead.region,
        status="",
        created_by=operator,
        created_at=_now(),
    )
    db.add(project)
    db.flush()

    # 3) 更新线索状态
    db_lead.status = "已转化"

    # 4) 创建活动记录
    activity = DBLeadActivity(
        id=str(uuid.uuid4()),
        lead_id=db_lead.id,
        activity_type="system",
        content=f"📌 线索转化: 已转化为客户「{customer.company}」并生成项目台账",
        operator=operator,
        created_at=_now(),
    )
    db.add(activity)
    db.commit()

    add_audit_log(
        db, operator, current_user.get("role", "商务负责人"),
        "线索转化",
        f"线索「{db_lead.title}」已转化为客户「{customer.company}」并生成项目台账",
        db_lead.id,
    )

    return {
        "customer": {
            "id": customer.id,
            "company": customer.company,
            "industry": customer.industry,
            "region": customer.region,
            "contactName": customer.contact_name,
            "contactPhone": customer.contact_phone,
            "sourceDb": customer.source_db,
            "budgetAmount": float(customer.budget_amount) if customer.budget_amount is not None else None,
            "requirements": customer.requirements,
            "createdAt": _dt_str(customer.created_at) if customer.created_at else "",
        },
        "project": {
            "id": project.id,
            "projectName": project.project_name,
            "projectAmount": float(project.project_amount) if project.project_amount is not None else None,
            "projectStage": project.project_stage,
            "region": project.region,
            "status": project.status,
            "createdAt": _dt_str(project.created_at) if project.created_at else "",
        },
        "activity": {
            "id": activity.id,
            "content": activity.content,
            "operator": activity.operator,
            "createdAt": _dt_str(activity.created_at) if activity.created_at else "",
        },
    }

# --- 数据源健康监控 (4.1.2) ---
def _default_push_targets() -> str:
    return json.dumps([{"receiverName": "商务专员A", "receiverRole": "商务专员", "channels": ["system", "email"]}])

def _configured_monitor_sources(db: Session) -> set:
    """汇总所有网监规则中当前配置的数据源"""
    sources: set = set()
    for c in db.query(DBSearchConfig).all():
        try:
            sources.update(json.loads(c.monitor_sources or "[]"))
        except Exception:
            pass
    return sources


def _sync_source_health_records(db: Session) -> None:
    """与网监配置对齐：移除已删除数据源的健康记录，并为新数据源创建记录"""
    sources = _configured_monitor_sources(db)
    for rec in db.query(DBSourceHealth).all():
        if rec.source not in sources:
            db.delete(rec)
    for s in sources:
        _get_or_create_source_health(db, s)
    db.commit()


SOURCE_HEALTH_ABNORMAL_THRESHOLD = 3


def _get_or_create_source_health(db: Session, source: str) -> DBSourceHealth:
    rec = db.query(DBSourceHealth).filter(DBSourceHealth.source == source).first()
    if not rec:
        rec = DBSourceHealth(
            source=source,
            status="healthy",
            consecutive_failures=0,
            total_fetches=0,
            total_failures=0,
        )
        db.add(rec)
        db.flush()
    return rec


def _record_source_success(db: Session, source: str, structure_valid: bool = True) -> None:
    if source not in _configured_monitor_sources(db):
        return
    rec = _get_or_create_source_health(db, source)
    rec.last_success_at = _now()
    rec.consecutive_failures = 0
    rec.total_fetches = (rec.total_fetches or 0) + 1
    rec.status = "healthy"
    rec.last_error_message = None
    rec.structure_change_detected = 0
    db.flush()


def _begin_source_fetch_session(db: Session, source: str) -> None:
    """新一轮抓取开始，连续失败仅在本次抓取内计数"""
    if source not in _configured_monitor_sources(db):
        return
    rec = _get_or_create_source_health(db, source)
    rec.consecutive_failures = 0
    db.flush()


def _record_source_attempt_failure(
    db: Session,
    source: str,
    error_message: str,
    attempt: int,
) -> None:
    """单次抓取内的第 N 次尝试失败（N=1,2,3）"""
    if source not in _configured_monitor_sources(db):
        return
    rec = _get_or_create_source_health(db, source)
    rec.consecutive_failures = attempt
    rec.last_failure_at = _now()
    rec.last_error_message = error_message
    if attempt >= SOURCE_FETCH_MAX_ATTEMPTS:
        rec.total_failures = (rec.total_failures or 0) + 1
    if attempt >= SOURCE_HEALTH_ABNORMAL_THRESHOLD:
        rec.status = "abnormal"
    else:
        rec.status = "degraded"
    db.flush()


def _record_source_failure(
    db: Session,
    source: str,
    error_message: str,
    structure_change: bool = False,
) -> None:
    """单次失败（如未配置地址，不重试）"""
    if source not in _configured_monitor_sources(db):
        return
    rec = _get_or_create_source_health(db, source)
    rec.consecutive_failures = (rec.consecutive_failures or 0) + 1
    rec.total_failures = (rec.total_failures or 0) + 1
    rec.last_failure_at = _now()
    rec.last_error_message = error_message
    if structure_change:
        rec.structure_change_detected = 1
    if rec.consecutive_failures >= SOURCE_HEALTH_ABNORMAL_THRESHOLD:
        rec.status = "abnormal"
    else:
        rec.status = "degraded"
    db.flush()


def _seed_source_health(db: Session) -> None:
    """兼容旧调用：同步数据源健康记录"""
    _sync_source_health_records(db)

def _seed_push_rules(db: Session) -> None:
    """如果推送规则表为空，创建默认规则"""
    count = db.query(DBPushRule).count()
    if count > 0:
        return
    rule = DBPushRule(
        id="push-rule-default",
        rule_name="默认推送规则",
        status="active",
        high_intent_mode="realtime",
        medium_intent_schedule_hour=18,
        medium_intent_schedule_minute=0,
        low_intent_silent=1,
        targets=_default_push_targets(),
    )
    db.add(rule)
    db.commit()

def _last_event_is_success(rec: DBSourceHealth) -> bool:
    if not rec.last_success_at:
        return False
    if not rec.last_failure_at:
        return True
    return rec.last_success_at >= rec.last_failure_at


def _source_health_row_to_api(r: DBSourceHealth) -> dict:
    show_alerts = not _last_event_is_success(r)
    return {
        "source": r.source,
        "status": "healthy" if _last_event_is_success(r) else r.status,
        "consecutiveFailures": 0 if _last_event_is_success(r) else (r.consecutive_failures or 0),
        "totalFetches": r.total_fetches,
        "totalFailures": r.total_failures,
        "structureChangeDetected": bool(r.structure_change_detected) if show_alerts else False,
        "lastSuccessAt": _dt_str(r.last_success_at) if r.last_success_at else None,
        "lastFailureAt": _dt_str(r.last_failure_at) if r.last_failure_at else None,
        "lastErrorMessage": r.last_error_message if show_alerts else None,
    }


def _align_source_health_records(db: Session, allowed: set) -> list:
    """按给定数据源列表对齐健康记录（仅操作 tbl_source_health）"""
    for rec in db.query(DBSourceHealth).all():
        if rec.source not in allowed:
            db.delete(rec)
    for s in allowed:
        _get_or_create_source_health(db, s)
    db.commit()
    if not allowed:
        return []
    rows = db.query(DBSourceHealth).filter(DBSourceHealth.source.in_(allowed)).all()
    return [_source_health_row_to_api(r) for r in rows]


class SourceHealthSyncRequest(BaseModel):
    sources: List[str] = []


@app.get("/api/source-health")
def get_source_health(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    allowed = _configured_monitor_sources(db)
    return _align_source_health_records(db, allowed)


@app.post("/api/source-health/sync")
def sync_source_health(
    body: SourceHealthSyncRequest,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    allowed = set(body.sources or [])
    return _align_source_health_records(db, allowed)

@app.post("/api/source-health/{source}/reset")
def reset_source_health(source: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role != "超级管理员":
        raise HTTPException(status_code=403, detail="仅超级管理员有权重置数据源状态")
    rec = db.query(DBSourceHealth).filter(DBSourceHealth.source == source).first()
    if not rec:
        rec = DBSourceHealth(source=source)
        db.add(rec)
    rec.status = "healthy"
    rec.consecutive_failures = 0
    rec.total_fetches = 0
    rec.total_failures = 0
    rec.structure_change_detected = 0
    rec.last_error_message = None
    db.commit()
    return {"success": True}

# --- 推送规则配置 (4.2.3) ---
@app.get("/api/push-rules")
def get_push_rules(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    _seed_push_rules(db)
    rows = db.query(DBPushRule).all()
    return [{
        "id": r.id,
        "ruleName": r.rule_name,
        "status": r.status,
        "highIntentMode": r.high_intent_mode,
        "mediumIntentScheduleHour": r.medium_intent_schedule_hour,
        "mediumIntentScheduleMinute": r.medium_intent_schedule_minute,
        "lowIntentSilent": bool(r.low_intent_silent),
        "targets": json.loads(r.targets) if r.targets else [],
    } for r in rows]

@app.put("/api/push-rules")
def update_push_rules(payload: dict, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权修改推送规则")
    rows = db.query(DBPushRule).all()
    if rows:
        rule = rows[0]
    else:
        rule = DBPushRule(id=str(uuid.uuid4()))
        db.add(rule)
    rule.rule_name = payload.get("ruleName", rule.rule_name)
    rule.status = payload.get("status", rule.status)
    rule.high_intent_mode = payload.get("highIntentMode", rule.high_intent_mode)
    rule.medium_intent_schedule_hour = payload.get("mediumIntentScheduleHour", rule.medium_intent_schedule_hour)
    rule.medium_intent_schedule_minute = payload.get("mediumIntentScheduleMinute", rule.medium_intent_schedule_minute)
    rule.low_intent_silent = 1 if payload.get("lowIntentSilent", True) else 0
    rule.targets = json.dumps(payload.get("targets", json.loads(rule.targets or "[]")), ensure_ascii=False)
    db.commit()
    return {"success": True}

@app.get("/api/push-records")
def get_push_records(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    rows = db.query(DBPushRecord).order_by(DBPushRecord.created_at.desc()).limit(50).all()
    return [{
        "id": r.id,
        "title": r.title,
        "receiverName": r.receiver_name,
        "channels": r.channels,
        "status": r.status,
        "retryCount": r.retry_count,
        "maxRetry": r.max_retry,
        "errorMessage": r.error_message,
        "content": r.content,
        "createdAt": _dt_str(r.created_at) if r.created_at else "",
    } for r in rows]

@app.post("/api/push/dispatch")
def dispatch_push(current_user: dict = Depends(get_current_user), db: Session = Depends(get_db)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权执行推送")
    # 查找A级(高意向)与B级(中意向)的未转化线索作模拟推送
    a_leads = db.query(DBLead).filter(DBLead.grade == "A", DBLead.status.notin_(["已转化", "已无效", "已流失"])).count()
    b_leads = db.query(DBLead).filter(DBLead.grade == "B", DBLead.status.notin_(["已转化", "已无效", "已流失"])).count()
    # 创建一条推送记录作为演示
    if a_leads > 0 or b_leads > 0:
        record = DBPushRecord(
            id=str(uuid.uuid4()),
            title=f"推送简报: {a_leads}条高意向+{b_leads}条中意向",
            receiver_name=current_user.get("name", "管理员"),
            channels="system",
            status="sent",
            content=f"高意向线索 {a_leads} 条, 中意向线索 {b_leads} 条",
        )
        db.add(record)
        db.commit()
    return {"highIntentCount": a_leads, "mediumIntentCount": b_leads}

# --- 客户档案与项目台账 (4.2.4) ---
@app.get("/api/customers")
def get_customers(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    rows = db.query(DBCustomerProfile).order_by(DBCustomerProfile.created_at.desc()).all()
    return [{
        "id": r.id,
        "company": r.company,
        "industry": r.industry,
        "region": r.region,
        "contactName": r.contact_name,
        "contactPhone": r.contact_phone,
        "sourceDb": r.source_db,
        "budgetAmount": float(r.budget_amount) if r.budget_amount is not None else None,
        "requirements": r.requirements,
        "convertedBy": r.converted_by or '',
        "convertedAt": _dt_str(r.converted_at) if r.converted_at else (_dt_str(r.created_at) if r.created_at else ''),
        "createdAt": _dt_str(r.created_at) if r.created_at else "",
    } for r in rows]

@app.get("/api/projects")
def get_projects(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    rows = db.query(DBProjectLedger).order_by(DBProjectLedger.created_at.desc()).all()
    return [{
        "id": r.id,
        "projectName": r.project_name,
        "projectAmount": float(r.project_amount) if r.project_amount is not None else None,
        "projectStage": r.project_stage,
        "region": r.region,
        "createdBy": r.created_by or '',
        "createdAt": _dt_str(r.created_at) if r.created_at else "",
    } for r in rows]

@app.get("/api/rules")
def get_rules(db: Session = Depends(get_db)):
    rules = get_or_create_scoring_rule(db)
    return rule_to_api(rules)

@app.put("/api/rules")
def update_rules(payload: RuleUpdate, db: Session = Depends(get_db)):
    rules = db.query(DBScoringRule).first()
    if not rules:
        rules = DBScoringRule()
        db.add(rules)
        db.commit()
        db.refresh(rules)

    rules.budget_weight = payload.budgetWeight
    rules.match_weight = payload.matchWeight
    rules.stage_weight = payload.stageWeight
    rules.qualification_weight = payload.qualificationWeight
    rules.region_weight = payload.regionWeight
    rules.warning_threshold = payload.warningThreshold
    db.commit()
    db.refresh(rules)

    all_leads = db.query(DBLead).yield_per(500)
    for lead in all_leads:
        total, grade = calculate_score(
            rules,
            int(lead.budget_score or 0),
            int(lead.match_score or 0),
            int(lead.stage_score or 0),
            int(lead.qualification_score or 0),
            int(lead.region_score or 0),
        )
        lead.total_score = total
        lead.grade = grade
    db.commit()

    add_audit_log(
        db,
        payload.operatorName or "操作员",
        payload.operatorRole or "系统管理员",
        "更新打分规则",
        "更新了系统5维打分权重，并触发全量重算。",
    )

    return {"success": True, "rule": rule_to_api(rules)}

@app.get("/api/search-configs", response_model=List[SearchConfigResponse])
def get_search_configs(db: Session = Depends(get_db)):
    configs = db.query(DBSearchConfig).all()
    if not configs:
        config = DBSearchConfig()
        db.add(config)
        db.commit()
        db.refresh(config)
        configs = [config]
    return [SearchConfigResponse(**search_config_to_api(c)) for c in configs]

@app.post("/api/search-config", response_model=dict)
def create_search_config(config_data: SearchConfigUpdate, db: Session = Depends(get_db)):
    config = DBSearchConfig(
        rule_name=config_data.ruleName,
        status=config_data.status,
        monitor_sources=json.dumps(config_data.monitorSources, ensure_ascii=False),
        keywords=json.dumps(config_data.keywords, ensure_ascii=False),
        exclude_keywords=json.dumps(config_data.excludeKeywords, ensure_ascii=False),
        regions=json.dumps(config_data.regions, ensure_ascii=False),
        industries=json.dumps(config_data.industries, ensure_ascii=False),
        categories=json.dumps(config_data.categories, ensure_ascii=False),
        frequency=config_data.frequency,
        budget_min=config_data.budgetMin,
        budget_max=config_data.budgetMax,
        purchase_types=json.dumps(config_data.purchaseTypes, ensure_ascii=False),
        time_range=config_data.timeRange,
    )
    db.add(config)
    db.commit()
    db.refresh(config)
    
    add_audit_log(
        db,
        config_data.operatorName or "操作员",
        config_data.operatorRole or "系统管理员",
        "新增网监规则",
        f"新增了网监规则: {config.rule_name}"
    )
    
    return {"success": True, "searchConfig": search_config_to_api(config)}

@app.put("/api/search-config/{config_id}", response_model=dict)
def update_search_config(config_id: int, config_data: SearchConfigUpdate, db: Session = Depends(get_db)):
    config = db.query(DBSearchConfig).filter(DBSearchConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
        
    config.rule_name = config_data.ruleName
    config.status = config_data.status
    config.monitor_sources = json.dumps(config_data.monitorSources, ensure_ascii=False)
    config.keywords = json.dumps(config_data.keywords, ensure_ascii=False)
    config.exclude_keywords = json.dumps(config_data.excludeKeywords, ensure_ascii=False)
    config.regions = json.dumps(config_data.regions, ensure_ascii=False)
    config.industries = json.dumps(config_data.industries, ensure_ascii=False)
    config.categories = json.dumps(config_data.categories, ensure_ascii=False)
    config.frequency = config_data.frequency
    config.budget_min = config_data.budgetMin
    config.budget_max = config_data.budgetMax
    config.purchase_types = json.dumps(config_data.purchaseTypes, ensure_ascii=False)
    config.time_range = config_data.timeRange

    db.commit()
    db.refresh(config)

    add_audit_log(
        db,
        config_data.operatorName or "操作员",
        config_data.operatorRole or "系统管理员",
        "更新网监规则",
        f"调整了网监监控配置: {config.rule_name}。"
    )
    
    return {"success": True, "searchConfig": search_config_to_api(config)}


@app.post("/api/search-configs", response_model=SearchConfigResponse)
def create_search_config_plural(config_data: SearchConfigUpdate, db: Session = Depends(get_db)):
    result = create_search_config(config_data, db)
    return SearchConfigResponse(**result["searchConfig"])


@app.put("/api/search-configs/{config_id}", response_model=SearchConfigResponse)
def update_search_config_plural(config_id: int, config_data: SearchConfigUpdate, db: Session = Depends(get_db)):
    result = update_search_config(config_id, config_data, db)
    return SearchConfigResponse(**result["searchConfig"])


@app.delete("/api/search-config/{config_id}")
def delete_search_config(config_id: int, db: Session = Depends(get_db)):
    config = db.query(DBSearchConfig).filter(DBSearchConfig.id == config_id).first()
    if not config:
        raise HTTPException(status_code=404, detail="Config not found")
    db.delete(config)
    db.commit()
    return {"success": True}

@app.get("/api/search-config/stats")
def get_search_config_stats(db: Session = Depends(get_db)):
    search_rules = search_config_to_api(db.query(DBSearchConfig).first())
    keywords = search_rules["keywords"]
    exclude_keywords = search_rules["excludeKeywords"]
    
    matched_count = 0
    if keywords:
        lead_rows = db.query(DBLead.title, DBLead.requirements, DBLead.industry).yield_per(500)
        for title, requirements, industry in lead_rows:
            text_to_search = f"{title or ''} {requirements or ''} {industry or ''}".lower()
            if any(kw.lower() in text_to_search for kw in keywords):
                if not any(ek.lower() in text_to_search for ek in exclude_keywords):
                    matched_count += 1
                
    return {
        "keywordCount": len(keywords),
        "matchedLeadsCount": matched_count
    }

def _lead_identity_key(title: str, company: str) -> str:
    clean_title = "".join((title or "").lower().split())
    clean_company = "".join((company or "").lower().split())
    return f"{clean_title}::{clean_company}"

def _pick(options: List[str], index: int, fallback: List[str]) -> str:
    values = options or fallback
    return values[index % len(values)]

def _is_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in ("http", "https") and bool(parsed.netloc)

def _looks_like_detail_url(url: str) -> bool:
    parsed = urlparse(url or "")
    query = parse_qs(parsed.query or "")
    return (
        "detail" in (parsed.path or "").lower()
        or "cid" in query
        or "name" in query
    )

def _title_hint_from_url(url: str) -> str:
    try:
        parsed = urlparse(url or "")
        query = parse_qs(parsed.query or "")
        name_values = query.get("name") or []
        if name_values:
            return _normalize_space(unquote(name_values[0]))
    except Exception:
        pass
    return ""

def _query_hint_from_url(url: str) -> str:
    try:
        parsed = urlparse(url or "")
        query = parse_qs(parsed.query or "")
        hints: List[str] = []
        for key in ("name", "key", "keyword", "kw"):
            for value in query.get(key) or []:
                normalized = _normalize_space(unquote(value))
                if normalized:
                    hints.append(normalized)
        return " ".join(hints)
    except Exception:
        return ""

def resolve_source_urls(sources: List[str]) -> List[Tuple[str, str]]:
    resolved: List[Tuple[str, str]] = []
    for source in sources:
        if _is_url(source):
            resolved.append((source, source))
            continue
        for url in SOURCE_URL_MAP.get(source, []):
            resolved.append((source, url))
    return resolved

def fetch_html(url: str, run_id: str) -> str:
    started_at = time.perf_counter()
    crawler_logger.info("run_id=%s stage=http_fetch_start url=%s", run_id, url)

    def _request_once(target_url: str, verify_ssl: bool, attempt: str):
        req_started_at = time.perf_counter()
        crawler_logger.info(
            "run_id=%s stage=http_fetch_attempt attempt=%s url=%s verify_ssl=%s",
            run_id,
            attempt,
            target_url,
            verify_ssl,
        )
        response = requests.get(
            target_url,
            headers=DEFAULT_HTTP_HEADERS,
            timeout=CRAWLER_REQUEST_TIMEOUT,
            verify=verify_ssl,
        )
        response.raise_for_status()
        if not response.encoding or response.encoding.lower() == "iso-8859-1":
            response.encoding = response.apparent_encoding or "utf-8"
        crawler_logger.info(
            "run_id=%s stage=http_fetch_attempt_done attempt=%s url=%s status=%s elapsed_ms=%s bytes=%s",
            run_id,
            attempt,
            target_url,
            response.status_code,
            int((time.perf_counter() - req_started_at) * 1000),
            len(response.content or b""),
        )
        return response.text

    # 尝试 1：严格 HTTPS/HTTP 请求
    try:
        html = _request_once(url, True, "strict")
        elapsed_ms = int((time.perf_counter() - started_at) * 1000)
        crawler_logger.info("run_id=%s stage=http_fetch_done url=%s elapsed_ms=%s", run_id, url, elapsed_ms)
        return html
    except requests.exceptions.SSLError as ssl_err:
        crawler_logger.warning(
            "run_id=%s stage=http_fetch_ssl_error url=%s error=%s",
            run_id,
            url,
            ssl_err,
        )

    # 尝试 2：允许不安全的 SSL（由环境变量控制）
    if CRAWLER_ALLOW_INSECURE_SSL:
        try:
            html = _request_once(url, False, "insecure_ssl")
            elapsed_ms = int((time.perf_counter() - started_at) * 1000)
            crawler_logger.info(
                "run_id=%s stage=http_fetch_done url=%s elapsed_ms=%s mode=insecure_ssl",
                run_id,
                url,
                elapsed_ms,
            )
            return html
        except Exception as insecure_err:
            crawler_logger.warning(
                "run_id=%s stage=http_fetch_insecure_failed url=%s error=%s",
                run_id,
                url,
                insecure_err,
            )

    # 尝试 3：如果原始 URL 是 HTTPS，回退到 HTTP
    parsed = urlparse(url)
    if CRAWLER_ALLOW_HTTP_FALLBACK and parsed.scheme == "https":
        http_url = parsed._replace(scheme="http").geturl()
        try:
            html = _request_once(http_url, True, "http_fallback")
            elapsed_ms = int((time.perf_counter() - started_at) * 1000)
            crawler_logger.info(
                "run_id=%s stage=http_fetch_done url=%s elapsed_ms=%s mode=http_fallback",
                run_id,
                http_url,
                elapsed_ms,
            )
            return html
        except Exception as http_err:
            crawler_logger.warning(
                "run_id=%s stage=http_fetch_http_fallback_failed url=%s error=%s",
                run_id,
                http_url,
                http_err,
            )

    raise requests.exceptions.SSLError(f"All fetch attempts failed for url={url}")


def fetch_html_with_retries(
    url: str,
    run_id: str,
    db: Optional[Session] = None,
    source: Optional[str] = None,
) -> str:
    """抓取失败当场自动重试；连续失败仅在本次抓取内按 1→2→3 计数"""
    if db and source:
        _begin_source_fetch_session(db, source)
    last_error: Optional[Exception] = None
    for attempt in range(1, SOURCE_FETCH_MAX_ATTEMPTS + 1):
        try:
            return fetch_html(url, run_id)
        except Exception as exc:
            last_error = exc
            if db and source:
                _record_source_attempt_failure(db, source, str(exc), attempt)
            crawler_logger.warning(
                "run_id=%s stage=source_fetch_retry attempt=%s/%s url=%s error=%s",
                run_id,
                attempt,
                SOURCE_FETCH_MAX_ATTEMPTS,
                url,
                exc,
            )
            if attempt < SOURCE_FETCH_MAX_ATTEMPTS:
                time.sleep(SOURCE_FETCH_RETRY_DELAY_SEC)
    raise last_error  # type: ignore[misc]

def _text_or_empty(node: Any) -> str:
    return node.get_text(" ", strip=True) if node else ""

def _normalize_space(value: str) -> str:
    return re.sub(r"\s+", " ", value or "").strip()

def _extract_date(text: str) -> Optional[str]:
    match = re.search(r"(20\d{2})[-年/.](\d{1,2})[-月/.](\d{1,2})", text or "")
    if not match:
        return None
    year, month, day = match.groups()
    return f"{int(year):04d}-{int(month):02d}-{int(day):02d}"


def _parse_ymd(value: str) -> Optional[date]:
    text = _normalize_space(value or "")
    if not text:
        return None
    extracted = _extract_date(text) or text[:10]
    try:
        return datetime.strptime(extracted[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def _resolve_time_range_bounds(time_range: str) -> Tuple[Optional[date], Optional[date]]:
    """
    解析监测配置中的 timeRange，返回闭区间 [start, end]（均为本地日历日）。
    支持：不限 / 近三天 / 近一周 / 近一月 / 近三月 / YYYY-MM-DD~YYYY-MM-DD / YYYY-MM-DD至YYYY-MM-DD
    """
    raw = _normalize_space(time_range or "") or "不限"
    if raw in ("不限", "全部", "无限制", "unlimited", "-"):
        return None, None

    today = datetime.now(timezone(timedelta(hours=8))).date()

    preset_days = {
        "近三天": 3,
        "近3天": 3,
        "近一周": 7,
        "近7天": 7,
        "近一月": 30,
        "近一个月": 30,
        "近30天": 30,
        "近三月": 90,
        "近三个月": 90,
        "近90天": 90,
    }
    if raw in preset_days:
        return today - timedelta(days=preset_days[raw] - 1), today

    custom = re.match(
        r"^(20\d{2}-\d{1,2}-\d{1,2})\s*[~～至\-—]\s*(20\d{2}-\d{1,2}-\d{1,2})$",
        raw,
    )
    if custom:
        start = _parse_ymd(custom.group(1))
        end = _parse_ymd(custom.group(2))
        if start and end and end < start:
            start, end = end, start
        return start, end

    # 单日当作当天
    single = _parse_ymd(raw)
    if single:
        return single, single
    return None, None


def _publish_date_in_range(publish_date: str, time_range: str) -> Tuple[bool, str]:
    start, end = _resolve_time_range_bounds(time_range)
    if start is None and end is None:
        return True, ""
    parsed = _parse_ymd(publish_date)
    if not parsed:
        # 有时间窗要求但解析不到日期时保守丢弃，避免脏数据入库
        return False, f"publish_date_missing timeRange={time_range}"
    if start and parsed < start:
        return False, f"publish_date_before_range:{parsed}<{start}"
    if end and parsed > end:
        return False, f"publish_date_after_range:{parsed}>{end}"
    return True, ""


def _time_range_to_sdt_edt(time_range: str) -> Tuple[str, str]:
    """供搜索框型 API（如 hbbid sdt/edt）使用的起止时间字符串。"""
    start, end = _resolve_time_range_bounds(time_range)
    if not start and not end:
        return "", ""
    sdt = f"{start.isoformat()} 00:00:00" if start else ""
    edt = f"{end.isoformat()} 23:59:59" if end else ""
    return sdt, edt

def _clamp_str(value: Optional[str], max_len: int) -> str:
    text = _normalize_space(value or "")
    return text[:max_len] if len(text) > max_len else text


REGION_NOISE_TOKENS = (
    "公告时间", "获取招标", "开标时间", "预算金额", "招标文件", "公告期限",
    "采购方式", "联系人", "联系电话", "项目名称", "采购人", "招标人",
    "每日上午", "每日下午", "获取招标文件的地点", "开标地点", "打印",
)


PROVINCES_STANDARD = (
    "北京市", "天津市", "河北省", "山西省", "内蒙古自治区",
    "辽宁省", "吉林省", "黑龙江省", "上海市", "江苏省",
    "浙江省", "安徽省", "福建省", "江西省", "山东省",
    "河南省", "湖北省", "湖南省", "广东省", "广西壮族自治区",
    "海南省", "重庆市", "四川省", "贵州省", "云南省",
    "西藏自治区", "陕西省", "甘肃省", "青海省", "宁夏回族自治区",
    "新疆维吾尔自治区", "台湾省", "香港特别行政区", "澳门特别行政区",
)

PROVINCE_ALIASES = {
    "新疆维吾尔自治区": "新疆维吾尔自治区", "新疆自治区": "新疆维吾尔自治区", "新疆": "新疆维吾尔自治区",
    "内蒙古自治区": "内蒙古自治区", "内蒙古": "内蒙古自治区", "内蒙": "内蒙古自治区",
    "广西壮族自治区": "广西壮族自治区", "广西自治区": "广西壮族自治区", "广西": "广西壮族自治区",
    "宁夏回族自治区": "宁夏回族自治区", "宁夏自治区": "宁夏回族自治区", "宁夏": "宁夏回族自治区",
    "西藏自治区": "西藏自治区", "西藏": "西藏自治区",
    "香港特别行政区": "香港特别行政区", "香港特区": "香港特别行政区", "香港": "香港特别行政区",
    "澳门特别行政区": "澳门特别行政区", "澳门特区": "澳门特别行政区", "澳门": "澳门特别行政区",
    "北京市": "北京市", "北京": "北京市",
    "上海市": "上海市", "上海": "上海市",
    "天津市": "天津市", "天津": "天津市",
    "重庆市": "重庆市", "重庆": "重庆市",
    "河北省": "河北省", "河北": "河北省",
    "山西省": "山西省", "山西": "山西省",
    "辽宁省": "辽宁省", "辽宁": "辽宁省",
    "吉林省": "吉林省", "吉林": "吉林省",
    "黑龙江省": "黑龙江省", "黑龙江": "黑龙江省",
    "江苏省": "江苏省", "江苏": "江苏省",
    "浙江省": "浙江省", "浙江": "浙江省",
    "安徽省": "安徽省", "安徽": "安徽省",
    "福建省": "福建省", "福建": "福建省",
    "江西省": "江西省", "江西": "江西省",
    "山东省": "山东省", "山东": "山东省",
    "河南省": "河南省", "河南": "河南省",
    "湖北省": "湖北省", "湖北": "湖北省",
    "湖南省": "湖南省", "湖南": "湖南省",
    "广东省": "广东省", "广东": "广东省",
    "海南省": "海南省", "海南": "海南省",
    "四川省": "四川省", "四川": "四川省",
    "贵州省": "贵州省", "贵州": "贵州省",
    "云南省": "云南省", "云南": "云南省",
    "陕西省": "陕西省", "陕西": "陕西省",
    "甘肃省": "甘肃省", "甘肃": "甘肃省",
    "青海省": "青海省", "青海": "青海省",
    "台湾省": "台湾省", "台湾": "台湾省",
}

CITY_TO_PROVINCE = {
    # 浙江省
    "杭州": "浙江省", "宁波": "浙江省", "温州": "浙江省", "嘉兴": "浙江省", "湖州": "浙江省",
    "绍兴": "浙江省", "金华": "浙江省", "衢州": "浙江省", "舟山": "浙江省", "台州": "浙江省", "丽水": "浙江省",
    "桐庐": "浙江省", "淳安": "浙江省", "建德": "浙江省", "萧山": "浙江省", "余杭": "浙江省", "临安": "浙江省", "富阳": "浙江省",
    "莼山": "浙江省", "象山": "浙江省", "宁海": "浙江省", "余姚": "浙江省", "慈溪": "浙江省", "奉化": "浙江省",
    "乐清": "浙江省", "瑞安": "浙江省", "永嘉": "浙江省", "平阳": "浙江省", "苍南": "浙江省", "文成": "浙江省", "泰顺": "浙江省",
    "德清": "浙江省", "长兴": "浙江省", "安吉": "浙江省", "海宁": "浙江省", "平湖": "浙江省", "桐乡": "浙江省",
    "诸暨": "浙江省", "嵊州": "浙江省", "新昌": "浙江省", "兰溪": "浙江省", "义乌": "浙江省", "东阳": "浙江省", "永康": "浙江省",
    "武义": "浙江省", "浦江": "浙江省", "磐安": "浙江省", "江山": "浙江省", "常山": "浙江省", "开化": "浙江省", "龙游": "浙江省",
    "临海": "浙江省", "温岭": "浙江省", "玉环": "浙江省", "天台": "浙江省", "仙居": "浙江省", "三门": "浙江省",
    "龙泉": "浙江省", "青田": "浙江省", "缙云": "浙江省", "遂昌": "浙江省", "松阳": "浙江省", "云和": "浙江省", "庆元": "浙江省", "景宁": "浙江省",
    # 广东省
    "广州": "广东省", "深圳": "广东省", "东莞": "广东省", "佛山": "广东省", "珠海": "广东省",
    "惠州": "广东省", "中山": "广东省", "江门": "广东省", "湛江": "广东省", "茂名": "广东省",
    "肇庆": "广东省", "汕头": "广东省", "潮州": "广东省", "揭阳": "广东省", "清远": "广东省",
    "韶关": "广东省", "梅州": "广东省", "汕尾": "广东省", "河源": "广东省", "阳江": "广东省", "云浮": "广东省",
    # 湖北省
    "武汉": "湖北省", "宜昌": "湖北省", "襄阳": "湖北省", "黄石": "湖北省", "十堰": "湖北省",
    "荆州": "湖北省", "鄂州": "湖北省", "荆门": "湖北省", "黄冈": "湖北省", "孝感": "湖北省",
    "咸宁": "湖北省", "随州": "湖北省", "恩施": "湖北省", "仙桃": "湖北省", "天门": "湖北省", "潜江": "湖北省",
    "神农架": "湖北省", "掇刀": "湖北省", "东宝": "湖北省", "沙洋": "湖北省", "钟祥": "湖北省", "京山": "湖北省",
    "孝南": "湖北省", "孝昌": "湖北省", "大悟": "湖北省", "云梦": "湖北省", "应城": "湖北省", "安陆": "湖北省", "汉川": "湖北省",
    "东山头": "湖北省", "红安": "湖北省", "太平桥": "湖北省", "毛岗岭": "湖北省", "麻城": "湖北省", "黄州": "湖北省", "团风": "湖北省",
    "罗田": "湖北省", "英山": "湖北省", "浠水": "湖北省", "蕲春": "湖北省", "黄梅": "湖北省", "武穴": "湖北省", "小池": "湖北省",
    "通城": "湖北省", "嘉鱼": "湖北省", "赤壁": "湖北省", "崇阳": "湖北省", "通山": "湖北省", "咸安": "湖北省",
    "曾都": "湖北省", "广水": "湖北省", "随县": "湖北省",
    "房县": "湖北省", "丹江口": "湖北省", "郧阳": "湖北省", "郧西": "湖北省", "竹山": "湖北省", "竹溪": "湖北省", "茅箭": "湖北省", "张湾": "湖北省",
    # 江苏省
    "南京": "江苏省", "苏州": "江苏省", "无锡": "江苏省", "常州": "江苏省", "南通": "江苏省",
    "徐州": "江苏省", "连云港": "江苏省", "淮安": "江苏省", "盐城": "江苏省", "扬州": "江苏省",
    "镇江": "江苏省", "泰州": "江苏省", "宿迁": "江苏省", "昆山": "江苏省", "江阴": "江苏省", "常熟": "江苏省",
    # 四川省
    "成都": "四川省", "绵阳": "四川省", "自贡": "四川省", "攀枝花": "四川省", "泸州": "四川省",
    "德阳": "四川省", "广元": "四川省", "遂宁": "四川省", "内江": "四川省", "乐山": "四川省",
    "南充": "四川省", "眉山": "四川省", "宜宾": "四川省", "广安": "四川省", "达州": "四川省",
    # 山东省
    "济南": "山东省", "青岛": "山东省", "淄博": "山东省", "枣庄": "山东省", "东营": "山东省",
    "烟台": "山东省", "潍坊": "山东省", "济宁": "山东省", "泰安": "山东省", "威海": "山东省",
    "日照": "山东省", "临沂": "山东省", "德州": "山东省", "聊城": "山东省", "滨州": "山东省", "菏泽": "山东省",
    # 河南省
    "郑州": "河南省", "开封": "河南省", "洛阳": "河南省", "平顶山": "河南省", "安阳": "河南省",
    "鹤壁": "河南省", "新乡": "河南省", "焦作": "河南省", "濮阳": "河南省", "许昌": "河南省",
    "漯河": "河南省", "三门峡": "河南省", "南阳": "河南省", "商丘": "河南省", "信阳": "河南省",
    "周口": "河南省", "驻马店": "河南省", "济源": "河南省",
    # 湖南省
    "长沙": "湖南省", "株洲": "湖南省", "湘潭": "湖南省", "衡阳": "湖南省", "邵阳": "湖南省",
    "岳阳": "湖南省", "常德": "湖南省", "张家界": "湖南省", "益阳": "湖南省", "郴州": "湖南省",
    # 福建省
    "福州": "福建省", "厦门": "福建省", "漳州": "福建省", "泉州": "福建省", "三明": "福建省",
    "莆田": "福建省", "南平": "福建省", "龙岩": "福建省", "宁德": "福建省", "晋江": "福建省",
    # 安徽省
    "合肥": "安徽省", "芜湖": "安徽省", "蚌埠": "安徽省", "淮南": "安徽省", "马鞍山": "安徽省",
    "淮北": "安徽省", "铜陵": "安徽省", "安庆": "安徽省", "黄山": "安徽省", "滁州": "安徽省",
    "阜阳": "安徽省", "宿州": "安徽省", "六安": "安徽省", "亳州": "安徽省", "池州": "安徽省", "宣城": "安徽省",
    # 江西省
    "南昌": "江西省", "景德镇": "江西省", "萍乡": "江西省", "九江": "江西省", "新余": "江西省",
    "鹰潭": "江西省", "赣州": "江西省", "吉安": "江西省", "宜春": "江西省", "抚州": "江西省", "上饶": "江西省",
    # 河北省
    "石家庄": "河北省", "唐山": "河北省", "秦皇岛": "河北省", "邯郸": "河北省", "邢台": "河北省",
    "保定": "河北省", "张家口": "河北省", "承德": "河北省", "沧州": "河北省", "廊坊": "河北省", "衡水": "河北省", "雄安": "河北省",
    # 陕西省
    "西安": "陕西省", "铜川": "陕西省", "宝鸡": "陕西省", "咸阳": "陕西省", "渭南": "陕西省",
    "延安": "陕西省", "汉中": "陕西省", "榆林": "陕西省", "安康": "陕西省", "商洛": "陕西省",
    # 贵州省
    "贵阳": "贵州省", "六盘水": "贵州省", "遵义": "贵州省", "安顺": "贵州省", "毕节": "贵州省", "铜仁": "贵州省",
    # 云南省
    "昆明": "云南省", "曲靖": "云南省", "玉溪": "云南省", "保山": "云南省", "昭通": "云南省",
    "丽江": "云南省", "普洱": "云南省", "临沧": "云南省", "楚雄": "云南省", "大理": "云南省",
    # 广西
    "南宁": "广西壮族自治区", "柳州": "广西壮族自治区", "桂林": "广西壮族自治区", "梧州": "广西壮族自治区",
    "北海": "广西壮族自治区", "防城港": "广西壮族自治区", "钦州": "广西壮族自治区", "贵港": "广西壮族自治区",
    "玉林": "广西壮族自治区", "百色": "广西壮族自治区", "贺州": "广西壮族自治区", "河池": "广西壮族自治区",
    # 新疆
    "乌鲁木齐": "新疆维吾尔自治区", "克拉玛依": "新疆维吾尔自治区", "吐鲁番": "新疆维吾尔自治区", "哈密": "新疆维吾尔自治区",
    "昌吉": "新疆维吾尔自治区", "博尔塔拉": "新疆维吾尔自治区", "巴音郭楞": "新疆维吾尔自治区", "阿克苏": "新疆维吾尔自治区",
    "克孜勒苏": "新疆维吾尔自治区", "喀什": "新疆维吾尔自治区", "和田": "新疆维吾尔自治区", "伊犁": "新疆维吾尔自治区",
    "塔城": "新疆维吾尔自治区", "阿勒泰": "新疆维吾尔自治区", "石河子": "新疆维吾尔自治区", "阿拉尔": "新疆维吾尔自治区",
    "图木舒克": "新疆维吾尔自治区", "五家渠": "新疆维吾尔自治区", "鄯善": "新疆维吾尔自治区",
    # 内蒙古
    "呼和浩特": "内蒙古自治区", "包头": "内蒙古自治区", "乌海": "内蒙古自治区", "赤峰": "内蒙古自治区",
    "通辽": "内蒙古自治区", "鄂尔多斯": "内蒙古自治区", "呼伦贝尔": "内蒙古自治区",
    # 东北
    "哈尔滨": "黑龙江省", "齐齐哈尔": "黑龙江省", "大庆": "黑龙江省",
    "长春": "吉林省", "吉林市": "吉林省", "延边": "吉林省",
    "沈阳": "辽宁省", "大连": "辽宁省", "鞍山": "辽宁省",
    # 山西
    "太原": "山西省", "大同": "山西省", "长治": "山西省",
    # 甘肃
    "兰州": "甘肃省", "天水": "甘肃省", "酒泉": "甘肃省",
    # 海南
    "海口": "海南省", "三亚": "海南省", "儋州": "海南省",
    # 宁夏
    "银川": "宁夏回族自治区", "石嘴山": "宁夏回族自治区", "吴忠": "宁夏回族自治区",
    # 青海
    "西宁": "青海省", "海东": "青海省",
    # 西藏
    "拉萨": "西藏自治区", "日喀则": "西藏自治区", "林芝": "西藏自治区",
}


LOCATION_STOP_KEYWORDS = (
    "建设项目规模", "建设规模", "项目建设规模", "项目规模", "工程规模", "主要规模",
    "项目建设内容", "建设内容", "工程内容", "主要建设内容", "建设主要内容", "项目概况",
    "主要招标范围", "主要招标内容", "招标范围", "标段划分", "采购需求", "合同履行",
    "计划工期", "工期要求", "工期", "质量要求", "出资比例", "资金来源", "出资方式", "资金落实",
    "投资估算", "总投资", "项目投资", "预算金额", "最高限价", "最高报价",
    "工程概况", "招标条件", "1.招标条件", "2.项目概况", "3.投标人", "3.1本次招标",
    "公告时间", "获取招标", "开标时间", "招标文件", "公告期限", "招标文件售价",
    "采购方式", "联系人", "联系电话", "项目名称", "采购人", "招标人", "建设单位", "代理机构",
    "每日上午", "每日下午", "获取招标文件的地点", "开标地点", "打印",
)


def _clean_location_text(value: str) -> str:
    raw = _normalize_space(value or "")
    if not raw or raw in ("待核验地区", "未知", "0"):
        return ""

    # 1. 强力剥离前缀引导词
    raw = re.sub(
        r"^(项目建设地点位于|项目地点位于|建设地点位于|工程建设地点位于|工程地点位于|"
        r"项目建设地点|项目建设地区|建设地点|项目地点|所在地区|行政区域|工程地点|项目位于|建设位于|位于|在)[：:\s]*",
        "",
        raw,
    )
    raw = re.sub(r"^(\d{4}年(?:度)?|[一二三四五六七八九十\d]+[、.：:]|关于|建设)\s*", "", raw)

    # 2. 截断后续拼接的混杂字段标签（如“建设规模：”、“建设项目规模：”等）
    for kw in LOCATION_STOP_KEYWORDS:
        idx = raw.find(kw)
        if idx >= 0:
            raw = raw[:idx].strip()

    # 3. 去除末尾标点符号和空格
    raw = re.sub(r"[。；;，,：:\s、]+$", "", raw).strip()
    return raw


def _sanitize_region(value: str) -> str:
    raw = _clean_location_text(value)
    if not raw:
        return ""

    # 1. 完整标准省/直辖市/自治区名
    for p in sorted(PROVINCES_STANDARD, key=len, reverse=True):
        if p in raw:
            # 提取如 "湖北省黄冈市红安县太平桥镇毛岗岭村" 或 "湖北省孝感市孝南区"
            if raw.startswith(p) and len(raw) <= 35:
                return raw
            m = re.search(rf"({re.escape(p)}[\u4e00-\u9fff]{{0,25}}?(?:市|区|县|旗|乡|镇|村|工业园(?:区)?|开发区)?)", raw)
            if m and len(m.group(1)) >= len(p):
                return m.group(1).rstrip("，,。；; ")
            return p

    # 2. 别名
    for alias, standard in sorted(PROVINCE_ALIASES.items(), key=lambda x: len(x[0]), reverse=True):
        if alias in raw:
            if len(raw) <= 30 and not any(kw in raw for kw in ("规模", "招标", "项目", "改造", "处理厂")):
                return raw
            return standard

    # 3. 市/区/县
    for city, province in sorted(CITY_TO_PROVINCE.items(), key=lambda x: len(x[0]), reverse=True):
        if city in raw:
            if len(raw) <= 30 and not any(kw in raw for kw in ("规模", "工艺", "招标", "改造", "处理厂")):
                return raw
            return province

    # 4. 正则提取省/自治区
    province = re.match(r"^([\u4e00-\u9fff]{2,15}(?:省|市|自治区|特别行政区))", raw)
    if province:
        cand = re.sub(r"^(年|年度|度|在|于|及|和|关于|建设)", "", province.group(1))
        if cand:
            return cand

    # 5. 常见行政区划特征词
    if re.search(r"(?:省|市|区|县|旗|乡|镇|村|工业园|园区|开发区|新区|街道|路|道|盟|州|大厦)", raw) and 2 <= len(raw) <= 30:
        return raw

    # 6. 非地理位置字符（如纯数字、单个汉字如“啊”、纯符号或无意义文本）判定为无效
    if len(raw) < 2 or re.match(r"^[\d\s\-_+=.,!@#$%^&*()，。！？、啊哦呃吧呢哈嘻呵测试]+$", raw):
        return ""

    if len(raw) <= 20 and not any(ch in raw for ch in ("1", "2", "3", "4", "5", "6", "7", "8", "9", "0")):
        return raw

    return ""


_COMPANY_NOISE_TOKENS = [
    "已经", "本次", "依法", "确定", "中标结果", "公告", "公示",
    "发布", "于", "在", "特此", "兹有", "如有", "有关",
    "评标结果", "中标人", "中标价", "招标编号",
    "需要说明", "其他事项", "说明的其他", "备注",
    "研究决定", "重新招标", "流标",
]

def _is_valid_company(name: str) -> bool:
    raw = _normalize_space(name or "")
    if not raw or len(raw) < 3:
        return False
    for token in _COMPANY_NOISE_TOKENS:
        if token in raw:
            return False
    if len(raw) > 25:
        return False
    return True


def _extract_region(text: str, title: str = "", company: str = "") -> str:
    raw = _extract_field(text, [
        "项目建设地点", "建设地点", "项目地点", "所在地区", "行政区域", "工程地点", "地区",
    ])
    sanitized = _sanitize_region(raw)
    if sanitized:
        return sanitized
    for candidate_text in (title, company, text):
        if not candidate_text:
            continue
        p = _sanitize_region(candidate_text)
        if p and (p in PROVINCES_STANDARD or any(p.startswith(prov) for prov in PROVINCES_STANDARD)):
            return p
    return ""


def _extract_field(text: str, labels: List[str]) -> str:
    stop_list = (
        r"[一二三四五六七八九十]+、|[0-9]+[.、]|"
        r"项目名称|招标项目|项目法人|招标人|采购人|采购单位|"
        r"项目批准|项目建设内容|建设内容|主要建设内容|工程内容|"
        r"项目建设地点|项目地点|建设地点|所在地区|行政区域|工程地点|"
        r"建设项目规模|建设规模|项目规模|工程规模|主要规模|规模|"
        r"投资估算|资金来源|主要招标范围|招标范围|标段划分|预计发布|"
        r"计划工期|工期要求|工期|工期日历天|质量要求|出资比例|出资方式|资金落实|"
        r"联系人|联系电话|电话|地址|统一社会信用代码|信用代码|组织机构代码|邮编|"
        r"代理机构|招标代理|传真|电子邮箱|邮箱|盖章|（盖章|\(盖章|"
        r"公告时间|获取招标|开标时间|预算金额|招标文件售价|"
        r"公告期限|采购方式|采购需求|合同履行|本项目|本公告|附件|"
        r"项目概况|工程概况|招标条件|1.招标条件|2.项目概况|3.投标人|$"
    )
    for label in labels:
        spaced_label = r"\s*".join(re.escape(c) for c in label)
        # strict: colon required
        strict = rf"{spaced_label}\s*[：:]\s*(.+?)(?=\s*(?:{stop_list}))"
        m = re.search(strict, text or "")
        if m:
            value = _normalize_space(m.group(1))
            return re.sub(r"[。；;，,]\s*$", "", value).strip()
        # loose: colon or whitespace required
        loose = rf"{spaced_label}(?:\s*[：:]|\s+)\s*(.+?)(?=\s*(?:{stop_list}))"
        m = re.search(loose, text or "")
        if m:
            value = _normalize_space(m.group(1))
            return re.sub(r"[。；;，,]\s*$", "", value).strip()
    return ""

# 湖北公共资源等站点常用「招标人 | 代理机构」双列联系方式，字段会重复出现：
#   联系人: 钱星宇 联系人: （空） 电话: 189... 电话: （空）
_CONTACT_FIELD_STOP = (
    "项目联系人", "联系人", "采购联系人", "经办人", "负责人",
    "项目联系电话", "联系电话", "采购单位联系方式", "代理机构联系方式", "电话",
    "招标人传真", "代理传真", "传真", "电子邮箱", "邮箱", "网址", "地址", "邮编",
    "开户银行", "账号", "账 号", "招标代理机构", "招标代理", "代理机构",
)
_CONTACT_NAME_REJECT = re.compile(
    r"(传真|电话|邮箱|地址|邮编|网址|银行|账号|代理|招标人|招标|机构|中心|公司|有限|"
    r"办事处|委员会|服务中心|建设|投资|政府)"
)


def _extract_contact_name(text: str) -> str:
    """从公告正文提取联系人姓名；兼容双列空字段与「联系人后不紧跟电话」的版式。"""
    labels = ["项目联系人", "联系人", "采购联系人", "经办人", "负责人"]
    stop = "|".join(re.escape(x) for x in _CONTACT_FIELD_STOP)
    for label in labels:
        pattern = (
            rf"{re.escape(label)}\s*[:：]?\s*"
            rf"([^\n\r，,。；;]{{1,40}}?)"
            rf"(?=(?:[；;，,。.\s]*)(?:{stop})|$)"
        )
        for match in re.finditer(pattern, text or ""):
            raw = _normalize_space(match.group(1))
            raw = re.sub(r"[：:]\s*$", "", raw).strip()
            if not raw:
                continue
            # 取开头中文姓名（含少数民族间隔号），避免吞进后续字段名
            name_match = re.match(r"^([\u4e00-\u9fff]{2,4}(?:[·•][\u4e00-\u9fff]{1,4})?)", raw)
            if not name_match:
                continue
            name = name_match.group(1)
            if _CONTACT_NAME_REJECT.search(name) or re.search(r"\d", name):
                continue
            return name
    return ""


def _extract_contact_phone(text: str) -> str:
    labels = ["项目联系电话", "联系电话", "采购单位联系方式", "代理机构联系方式", "电话"]
    for label in labels:
        pattern = rf"{re.escape(label)}\s*[:：]?\s*([0-9\-]{{7,20}})"
        for match in re.finditer(pattern, text or ""):
            phone = _normalize_space(match.group(1))
            if phone:
                return phone

    generic = re.findall(
        r"(?:\+?86[-\s]?)?(?:1[3-9]\d{9}|0\d{2,3}-?\d{7,8}|400[-\s]?\d{3}[-\s]?\d{4})",
        text or "",
    )
    if generic:
        return _normalize_space(generic[0])
    return ""

def _extract_budget(text: str) -> Optional[float]:
    prioritized_patterns = [
        r"(?:合同估算价|合同估算金额)[^0-9]{0,30}([0-9]+(?:\.[0-9]+)?)\s*(万元|元)",
        r"(?:招标控制价|最高限价)[^0-9]{0,30}([0-9]+(?:\.[0-9]+)?)\s*(万元|元)",
        r"(?:投资估算价|投资估算|项目估算价|估算价|总投资)[^0-9]{0,30}([0-9]+(?:\.[0-9]+)?)\s*(万元|元)",
        r"(?:预算金额|预算金额约)[^0-9]{0,30}([0-9]+(?:\.[0-9]+)?)\s*(万元|元)",
    ]
    for pattern in prioritized_patterns:
        for match in re.finditer(pattern, text or ""):
            # 跳过资格要求/业绩条款中的门槛条件，如"9000万元及以上"
            trailing = (text or "")[match.end():match.end() + 6]
            if any(token in trailing for token in ("及以上", "以上", "以下")):
                continue
            amount = float(match.group(1))
            unit = match.group(2)
            return round(amount / 10000, 2) if unit == "元" else amount
    return None


def _matches_dimension_filters(
    text: str,
    region: str,
    budget_wan: Optional[float],
    publish_date: str,
    search_rules: Dict[str, Any],
) -> Tuple[bool, str]:
    budget_min = _coerce_int(search_rules.get("budgetMin"), DEFAULT_SEARCH_CONFIG["budgetMin"])
    budget_max = _coerce_int(search_rules.get("budgetMax"), DEFAULT_SEARCH_CONFIG["budgetMax"])
    if budget_wan is not None and not (budget_min <= budget_wan <= budget_max):
        return False, f"budget_out_of_range:{budget_wan}"

    time_ok, time_reason = _publish_date_in_range(
        publish_date,
        str(search_rules.get("timeRange") or DEFAULT_SEARCH_CONFIG["timeRange"]),
    )
    if not time_ok:
        return False, time_reason

    # 地区限制：忽略「全部/不限/全国」占位；采购类型/分类多为展示维度，不作硬过滤以免误杀
    regions = [
        r for r in _coerce_list(search_rules.get("regions"), [])
        if r and r not in ("全部", "不限", "全国")
    ]
    if regions:
        haystack = f"{region or ''} {text or ''}".lower()
        if not any(r.lower() in haystack for r in regions):
            return False, f"region_not_matched:{'/'.join(regions)}"

    return True, ""

def _set_query_param(url: str, key: str, value: str) -> str:
    parsed = urlparse(url or "")
    pairs = [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k.lower() != key.lower()]
    pairs.append((key, str(value)))
    return urlunparse(parsed._replace(query=urlencode(pairs, doseq=True)))


def _detect_list_page_param(url: str) -> Tuple[Optional[str], Optional[int]]:
    """若 URL 已带分页参数，返回 (参数名, 当前页码)。"""
    query = parse_qs(urlparse(url or "").query or "", keep_blank_values=True)
    for key in LIST_PAGE_QUERY_KEYS:
        if key in query and query[key]:
            raw = (query[key][0] or "").strip()
            if raw.isdigit():
                return key, max(1, int(raw))
            return key, 1
    # 大小写不敏感再扫一遍
    lower_map = {k.lower(): k for k in query.keys()}
    for key in LIST_PAGE_QUERY_KEYS:
        real = lower_map.get(key.lower())
        if real and query.get(real):
            raw = (query[real][0] or "").strip()
            if raw.isdigit():
                return real, max(1, int(raw))
            return real, 1
    return None, None


def _expand_list_page_urls(base_url: str, max_pages: int) -> List[str]:
    """
    基于白名单原始 URL 生成候选列表页。
    - 始终包含原始 URL（兼容无分页站点 / 首页即列表）
    - 若已有 page/pageNo 等参数：按该参数翻 1..max_pages
    - 若无分页参数：保留原 URL 作为第 1 页，再追加 ?page=2..N
    """
    max_pages = max(1, int(max_pages or 1))
    urls: List[str] = []
    seen = set()

    def _add(u: str) -> None:
        if u and u not in seen:
            seen.add(u)
            urls.append(u)

    _add(base_url)
    if max_pages <= 1:
        return urls

    page_key, _current = _detect_list_page_param(base_url)
    if page_key:
        for page in range(1, max_pages + 1):
            _add(_set_query_param(base_url, page_key, str(page)))
    else:
        for page in range(2, max_pages + 1):
            _add(_set_query_param(base_url, "page", str(page)))

    return urls


def _discover_next_list_url(html: str, current_url: str) -> Optional[str]:
    """从列表页 HTML 中发现「下一页」链接（适配 path 翻页 / JS 站点失败时的兜底）。"""
    if not html:
        return None
    soup = BeautifulSoup(html, "html.parser")
    next_texts = ("下一页", "下页", "后一页", "next", "Next", "»", "›")
    for anchor in soup.find_all("a", href=True):
        text = _normalize_space(anchor.get_text(" ", strip=True))
        title = _normalize_space(anchor.get("title") or "")
        rel = _normalize_space(anchor.get("rel") or "").lower()
        href = anchor.get("href") or ""
        if href.startswith("#") or href.lower().startswith("javascript:"):
            continue
        matched = (
            text in next_texts
            or title in next_texts
            or "next" in rel
            or text.endswith("下一页")
            or "下一页" in text
        )
        if not matched:
            continue
        next_url = urljoin(current_url, href)
        if not _is_url(next_url):
            continue
        if next_url.rstrip("/") == (current_url or "").rstrip("/"):
            continue
        return next_url
    return None


def _is_noise_link(detail_url: str, title: str, base_url: str) -> bool:
    normalized_url = (detail_url or "").lower()
    normalized_title = (title or "").lower()
    if not normalized_url:
        return True

    if normalized_url.rstrip("/") == (base_url or "").lower().rstrip("/"):
        return True

    noise_url_tokens = (
        "login",
        "sso",
        "system_service",
        "web-enterprise",
        "javascript:",
    )
    if any(token in normalized_url for token in noise_url_tokens):
        return True

    # 采购页面常见的公告/项目特征词。
    useful_url_tokens = (
        "gkzb",
        "jyfw",
        "zfcg",
        "gg",
        "notice",
        "detail",
        "bid",
        "project",
        "t20",
    )
    useful_title_tokens = (
        "招标",
        "采购",
        "公告",
        "项目",
        "工程",
        "污水",
        "改造",
    )

    if not any(token in normalized_url for token in useful_url_tokens) and not any(token in title for token in useful_title_tokens):
        return True
    if not any(token in title for token in useful_title_tokens) and "gkzb" not in normalized_url:
        return True
    return False

def parse_notice_links(
    html: str,
    base_url: str,
    source_name: str,
    run_id: str,
    seen_urls: Optional[set] = None,
    limit: Optional[int] = None,
) -> List[Dict[str, Any]]:
    soup = BeautifulSoup(html, "html.parser")
    notices: List[Dict[str, Any]] = []
    seen = seen_urls if seen_urls is not None else set()
    max_count = limit if limit is not None else CRAWLER_MAX_LINKS_PER_SOURCE

    for anchor in soup.find_all("a", href=True):
        if max_count <= 0:
            break
        title = _normalize_space(anchor.get_text(" ", strip=True))
        if len(title) < 6:
            continue
        href = anchor.get("href", "")
        detail_url = urljoin(base_url, href)
        if not _is_url(detail_url) or detail_url in seen:
            continue
        if detail_url.lower().endswith((".jpg", ".png", ".gif", ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".zip")):
            continue
        if _is_noise_link(detail_url, title, base_url):
            crawler_logger.info(
                "run_id=%s stage=parse_list_skip_noise source=%s detail_url=%s title=%s",
                run_id,
                source_name,
                detail_url,
                title,
            )
            continue

        parent_text = _normalize_space(_text_or_empty(anchor.find_parent(["li", "tr", "div", "p"])) or title)
        notices.append({
            "title": title,
            "detailUrl": detail_url,
            "sourceDb": source_name,
            "listSummary": parent_text,
            "publishDate": _extract_date(parent_text),
        })
        crawler_logger.info(
            "run_id=%s stage=parse_list_url source=%s detail_url=%s",
            run_id,
            source_name,
            detail_url,
        )
        seen.add(detail_url)

        if len(notices) >= max_count:
            break

    crawler_logger.info(
        "run_id=%s stage=parse_list source=%s url=%s links=%s",
        run_id,
        source_name,
        base_url,
        len(notices),
    )
    return notices


def _strip_html_text(value: str) -> str:
    if not value:
        return ""
    return _normalize_space(BeautifulSoup(value, "html.parser").get_text(" ", strip=True))


def _is_hbbid_fullsearch_url(url: str) -> bool:
    """湖北云平台 fullsearch：接口一次返回全量，前端再分页，勿做 ?page= HTML 翻页。"""
    parsed = urlparse(url or "")
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()
    return "hbbidcloud.cn" in host and "fullsearch" in path


def _is_full_result_list_url(url: str) -> bool:
    """一次返回全量结果集的列表入口（当前：hbbidcloud fullsearch）。"""
    return _is_hbbid_fullsearch_url(url)


def _hbbid_yycx_from_cnum(cnum: str) -> str:
    # 与 fullsearch.js CNUM_YYCX_MAP 对齐的常用映射；未知则走省级 hubei
    mapping = {
        "001": "hubei",
        "006": "shengbenji",
        "014": "huanggang",
        "018": "zhushan",
        "023": "shiyan",
    }
    return mapping.get((cnum or "").strip(), "hubei")


def _resolve_hbbid_detail_url(linkurl: str, cnum: str) -> str:
    link = (linkurl or "").strip()
    if not link:
        return ""
    if _is_url(link):
        return link
    yycx = _hbbid_yycx_from_cnum(cnum)
    # 站点 rewriteLinks: "/" + yycx + originalLink
    if not link.startswith("/"):
        link = "/" + link
    rewritten = f"/{yycx}{link}"
    return urljoin(HBBID_FULLSEARCH_BASE + "/", rewritten.lstrip("/"))


def _hbbid_wd_from_url(source_url: str) -> str:
    query = parse_qs(urlparse(source_url or "").query or "", keep_blank_values=True)
    vals = query.get("wd") or []
    return unquote(vals[0]).strip() if vals and vals[0] is not None else ""


def _hbbid_search_terms(source_url: str, search_rules: Optional[Dict[str, Any]]) -> List[str]:
    """
    搜索框型源关键词策略：
    - 优先使用监测配置 keywords（黑白名单里的核心包含词）逐词查询；
    - 单次 API 的 wd 对中文分词/OR 组合不稳定，故按词分别请求后去重合并；
    - 最多 HBBID_MAX_KEYWORD_QUERIES 个，避免请求爆炸；
    - 无监控词时回退到白名单 URL 的 wd。
    """
    terms: List[str] = []
    seen = set()
    for kw in (search_rules or {}).get("keywords") or []:
        term = _normalize_space(str(kw))
        if not term or term.lower() in seen:
            continue
        seen.add(term.lower())
        terms.append(term)
        if len(terms) >= max(1, HBBID_MAX_KEYWORD_QUERIES):
            break
    if terms:
        return terms
    fallback = _hbbid_wd_from_url(source_url)
    return [fallback] if fallback else []


def _hbbid_exclude_wd(source_url: str, search_rules: Optional[Dict[str, Any]]) -> str:
    query = parse_qs(urlparse(source_url or "").query or "", keep_blank_values=True)
    url_exc = unquote((query.get("exc_wd") or [""])[0] or "").strip()
    excludes = [
        _normalize_space(str(x))
        for x in ((search_rules or {}).get("excludeKeywords") or [])
        if _normalize_space(str(x))
    ]
    # 监控排除词优先；URL exc_wd 作为补充
    merged: List[str] = []
    seen = set()
    for item in excludes + ([url_exc] if url_exc else []):
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        merged.append(item)
    return " ".join(merged)


def _post_hbbid_fullsearch(
    source_url: str,
    wd: str,
    cnum: str,
    inc_wd: str,
    exc_wd: str,
    sdt: str,
    edt: str,
    batch_rn: int,
    run_id: str,
    db: Optional[Session],
    source_name: str,
) -> Dict[str, Any]:
    payload = {
        "token": "",
        "pn": 0,
        "rn": batch_rn,
        "sdt": sdt or "",
        "edt": edt or "",
        "wd": quote(wd),
        "inc_wd": quote(inc_wd) if inc_wd else "",
        "exc_wd": quote(exc_wd) if exc_wd else "",
        "fields": "title;content",
        "cnum": cnum,
        "sort": '{"webdate":"0"}',
        "ssort": "title",
        "cl": 300,
        "terminal": "",
        "condition": None,
        "time": None,
        "highlights": "",
        "statistics": None,
        "unionCondition": None,
        "accuracy": "",
        "noParticiple": "0",
        "searchRange": None,
        "opCondition": "",
        "opType": 0,
    }

    last_error: Optional[Exception] = None
    for attempt in range(1, SOURCE_FETCH_MAX_ATTEMPTS + 1):
        try:
            response = requests.post(
                HBBID_FULLSEARCH_API,
                headers={
                    **DEFAULT_HTTP_HEADERS,
                    "Content-Type": "application/json;charset=UTF-8",
                    "Origin": HBBID_FULLSEARCH_BASE,
                    "Referer": source_url,
                    "Accept": "application/json, text/javascript, */*; q=0.01",
                },
                data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
                timeout=max(CRAWLER_REQUEST_TIMEOUT, 30),
            )
            response.raise_for_status()
            return response.json() if response.content else {}
        except Exception as exc:
            last_error = exc
            if db and source_name:
                _record_source_attempt_failure(db, source_name, str(exc), attempt)
            crawler_logger.warning(
                "run_id=%s stage=hbbid_fullsearch_retry attempt=%s/%s wd=%s error=%s",
                run_id,
                attempt,
                SOURCE_FETCH_MAX_ATTEMPTS,
                wd,
                exc,
            )
            if attempt < SOURCE_FETCH_MAX_ATTEMPTS:
                time.sleep(SOURCE_FETCH_RETRY_DELAY_SEC)
    raise last_error  # type: ignore[misc]


def _fetch_hbbid_fullsearch_notices(
    source_name: str,
    source_url: str,
    run_id: str,
    db: Optional[Session] = None,
    search_rules: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Dict[str, Any]], int, bool]:
    """
    调用 getFullTextDataNew，按监控关键词逐词拉取并去重合并（站点单次上限 500），
    再按 CRAWLER_MAX_LINKS_PER_SOURCE 截断。支持 sdt/edt 时间窗与 exc_wd 排除词。
    """
    query = parse_qs(urlparse(source_url).query or "", keep_blank_values=True)

    def _q(name: str, default: str = "") -> str:
        vals = query.get(name) or []
        return unquote(vals[0]) if vals and vals[0] is not None else default

    search_terms = _hbbid_search_terms(source_url, search_rules)
    if not search_terms:
        raise ValueError("hbbid fullsearch 缺少搜索词：请配置监测关键词或在白名单 URL 中提供 wd")

    cnum = (_q("cnum") or "001").strip()
    inc_wd = _q("inc_wd")
    exc_wd = _hbbid_exclude_wd(source_url, search_rules)
    time_range = str((search_rules or {}).get("timeRange") or DEFAULT_SEARCH_CONFIG["timeRange"])
    # URL 显式 sdt/edt 优先，否则由监测 timeRange 推导
    sdt = _q("sdt")
    edt = _q("edt")
    if not sdt and not edt:
        sdt, edt = _time_range_to_sdt_edt(time_range)

    max_links = max(1, CRAWLER_MAX_LINKS_PER_SOURCE)
    batch_rn = min(HBBID_FULLSEARCH_BATCH_RN, max_links)

    crawler_logger.info(
        "run_id=%s stage=hbbid_fullsearch_start source=%s terms=%s cnum=%s sdt=%s edt=%s rn=%s",
        run_id,
        source_name,
        search_terms,
        cnum,
        sdt,
        edt,
        batch_rn,
    )
    if db and source_name:
        _begin_source_fetch_session(db, source_name)

    notices: List[Dict[str, Any]] = []
    seen: set = set()
    api_total = 0

    for term in search_terms:
        if len(notices) >= max_links:
            break
        data = _post_hbbid_fullsearch(
            source_url=source_url,
            wd=term,
            cnum=cnum,
            inc_wd=inc_wd,
            exc_wd=exc_wd,
            sdt=sdt,
            edt=edt,
            batch_rn=batch_rn,
            run_id=run_id,
            db=db,
            source_name=source_name,
        )
        result = data.get("result") if isinstance(data, dict) else None
        records = (result or {}).get("records") if isinstance(result, dict) else None
        if not isinstance(records, list):
            records = []
        totalcount = (result or {}).get("totalcount") if isinstance(result, dict) else len(records)
        try:
            api_total += int(totalcount or 0)
        except (TypeError, ValueError):
            api_total += len(records)

        for rec in records:
            if not isinstance(rec, dict):
                continue
            title = _strip_html_text(str(rec.get("title") or ""))
            content = _strip_html_text(str(rec.get("content") or ""))
            detail_url = _resolve_hbbid_detail_url(str(rec.get("linkurl") or ""), cnum)
            if len(title) < 6 or not detail_url or detail_url in seen:
                continue
            if not _is_url(detail_url):
                continue
            date_text = str(rec.get("webdate") or rec.get("infodate") or "")
            publish_date = _extract_date(date_text) or _extract_date(content) or ""
            # 双保险：API sdt/edt 之外再按配置时间窗过滤
            in_range, _reason = _publish_date_in_range(publish_date, time_range)
            if not in_range:
                continue
            seen.add(detail_url)
            notices.append({
                "title": title,
                "detailUrl": detail_url,
                "sourceDb": source_name,
                "listSummary": content or title,
                "publishDate": publish_date,
            })
            if len(notices) >= max_links:
                break

    crawler_logger.info(
        "run_id=%s stage=hbbid_fullsearch_done source=%s api_total=%s ingested=%s terms=%s",
        run_id,
        source_name,
        api_total,
        len(notices),
        search_terms,
    )
    return notices, len(search_terms), bool(notices)


def fetch_paginated_notice_links(
    source_name: str,
    source_url: str,
    run_id: str,
    db: Optional[Session] = None,
    search_rules: Optional[Dict[str, Any]] = None,
) -> Tuple[List[Dict[str, Any]], int, bool]:
    """
    对白名单列表 URL 做通用翻页抓取。
    返回: (公告列表, 成功抓取的列表页数, 是否解析到至少一条链接)
    兼容：无分页站点、?page=/pageNo 参数翻页、HTML「下一页」、
    以及一次返回全量结果的 fullsearch API（不做无意义的 ?page= 翻页）。
    """
    if _is_full_result_list_url(source_url):
        return _fetch_hbbid_fullsearch_notices(
            source_name, source_url, run_id, db=db, search_rules=search_rules
        )

    max_pages = max(1, CRAWLER_MAX_PAGES)
    max_links = max(1, CRAWLER_MAX_LINKS_PER_SOURCE)
    notices: List[Dict[str, Any]] = []
    seen_detail_urls: set = set()
    visited_list_urls: set = set()
    pages_ok = 0

    page_queue = _expand_list_page_urls(source_url, max_pages)
    page_idx = 0

    while page_queue and page_idx < max_pages and len(notices) < max_links:
        page_url = page_queue.pop(0)
        if page_url in visited_list_urls:
            continue
        visited_list_urls.add(page_url)
        page_idx += 1

        crawler_logger.info(
            "run_id=%s stage=list_page_fetch source=%s page=%s/%s list_url=%s",
            run_id,
            source_name,
            page_idx,
            max_pages,
            page_url,
        )
        try:
            # 仅首页计入数据源健康重试；后续页失败只停翻页，不污染健康统计
            if page_idx == 1:
                list_html = fetch_html_with_retries(page_url, run_id, db=db, source=source_name)
            else:
                list_html = fetch_html(page_url, run_id)
        except Exception as exc:
            crawler_logger.warning(
                "run_id=%s stage=list_page_fetch_failed source=%s page=%s url=%s error=%s",
                run_id,
                source_name,
                page_idx,
                page_url,
                exc,
            )
            # 首页失败则整源失败；后续页失败则停止翻页
            if page_idx == 1:
                raise
            break

        pages_ok += 1
        remaining = max_links - len(notices)
        page_notices = parse_notice_links(
            list_html,
            page_url,
            source_name,
            run_id,
            seen_urls=seen_detail_urls,
            limit=remaining,
        )
        notices.extend(page_notices)

        crawler_logger.info(
            "run_id=%s stage=list_page_done source=%s page=%s new_links=%s total_links=%s",
            run_id,
            source_name,
            page_idx,
            len(page_notices),
            len(notices),
        )

        # 本页无新增链接：停止翻页（避免空页继续扫）
        if not page_notices:
            break

        # HTML「下一页」发现：补充进队列（适配 path 翻页）
        if page_idx < max_pages and len(notices) < max_links:
            next_url = _discover_next_list_url(list_html, page_url)
            if next_url and next_url not in visited_list_urls and next_url not in page_queue:
                page_queue.append(next_url)
                crawler_logger.info(
                    "run_id=%s stage=list_next_discovered source=%s next_url=%s",
                    run_id,
                    source_name,
                    next_url,
                )

    return notices, pages_ok, bool(notices)

def parse_notice_detail(html: str) -> Dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    title = _normalize_space(_text_or_empty(soup.find(["h1", "h2"])))
    body_node = soup.find("article") or soup.find(class_=re.compile("(content|article|detail|main)", re.I)) or soup.body
    body_text = _normalize_space(_text_or_empty(body_node))
    company = _extract_field(body_text, [
        "项目法人（招标人）", "项目法人(招标人)", "采购人名称", "招标人名称", "采购单位名称",
        "采购人（甲方）", "采购人(甲方)",
        "采购人", "招标人", "采购单位", "建设单位", "项目业主", "甲方",
    ])

    return {
        "title": title,
        "body": body_text,
        "publishDate": _extract_date(body_text) or "",
        "company": company,
        "region": _extract_region(body_text, title=title, company=company),
    }

def _expand_keywords(keywords: List[str]) -> List[str]:
    expanded: List[str] = []
    for keyword in keywords or []:
        normalized = _normalize_space(keyword)
        if not normalized:
            continue
        expanded.append(normalized)
        for group in KEYWORD_SYNONYM_GROUPS:
            if any(term in normalized for term in group):
                expanded.extend(group)
    seen = set()
    return [kw for kw in expanded if not (kw in seen or seen.add(kw))]

def _matches_crawler_rules(text: str, keywords: List[str], exclude_keywords: List[str]) -> Tuple[bool, str]:
    lower_text = (text or "").lower()
    normalized_keywords = [_normalize_space(keyword) for keyword in keywords or [] if _normalize_space(keyword)]
    requires_water_domain = any(
        term.lower() in keyword.lower()
        for keyword in normalized_keywords
        for term in WATER_DOMAIN_TERMS
    )
    if requires_water_domain and not any(term.lower() in lower_text for term in WATER_DOMAIN_TERMS):
        return False, f"water_domain_not_matched keywords={keywords}"

    expanded_keywords = _expand_keywords(keywords)
    if expanded_keywords and not any(keyword.lower() in lower_text for keyword in expanded_keywords):
        return False, f"keyword_not_matched keywords={keywords} expanded={expanded_keywords}"
    blocked = next((keyword for keyword in exclude_keywords if keyword.lower() in lower_text), "")
    if blocked:
        return False, f"exclude_keyword:{blocked}"
    return True, ""

def _score_from_notice(text: str, budget_wan: Optional[float]) -> Dict[str, int]:
    budget_score = 60
    if budget_wan is not None:
        budget_score = 85 if budget_wan >= 1000 else 75 if budget_wan >= 300 else 65
    match_score = 82 if any(word in text for word in ("招标公告", "竞争性磋商", "询价公告", "采购公告")) else 65
    qualification_score = 84 if any(word in text for word in ("PLC", "自动化", "智能", "在线监测", "运维")) else 68
    stage_score = 45 if any(word in text for word in ("公开招标", "竞争性")) else 55
    region_score = 82 if any(word in text for word in ("广东", "广州", "深圳", "佛山", "东莞")) else 60
    return {
        "budgetScore": budget_score,
        "matchScore": match_score,
        "qualificationScore": qualification_score,
        "stageScore": stage_score,
        "regionScore": region_score,
    }

def collect_web_candidates(
    search_rules: Dict[str, Any],
    run_id: str,
    db: Optional[Session] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    keywords = search_rules["keywords"] or DEFAULT_CRAWLER_KEYWORDS
    exclude_keywords = search_rules["excludeKeywords"]
    sources = search_rules["monitorSources"] or DEFAULT_CRAWLER_SOURCES
    industries = search_rules["industries"] or DEFAULT_CRAWLER_INDUSTRIES
    categories = search_rules["categories"] or DEFAULT_CRAWLER_CATEGORIES
    source_urls = resolve_source_urls(sources)
    candidates: List[Dict[str, Any]] = []
    stats: Dict[str, Any] = {
        "mode": "web",
        "scanned": 0,
        "filtered": 0,
        "sourceCount": len(source_urls),
        "fetchErrors": 0,
        "unresolvedSources": [source for source in sources if not _is_url(source) and source not in SOURCE_URL_MAP],
    }

    if not source_urls:
        crawler_logger.warning("run_id=%s stage=source_resolve action=no_fetchable_urls sources=%s", run_id, sources)
        if db:
            for source in stats["unresolvedSources"]:
                _record_source_failure(db, source, "未配置抓取地址")
        return candidates, stats

    if db:
        for source in stats["unresolvedSources"]:
            _record_source_failure(db, source, "未配置抓取地址")

    crawler_logger.info(
        "run_id=%s stage=source_resolve action=resolved source_urls=%s",
        run_id,
        [url for _, url in source_urls],
    )

    for source_name, source_url in source_urls:
        crawler_logger.info(
            "run_id=%s stage=source_scan_start source=%s list_url=%s",
            run_id,
            source_name,
            source_url,
        )

        # 部分监控源是详情页 URL（如 business/detail?cid=...）。
        # 将其作为单条候选项页面而非列表页处理。
        if _looks_like_detail_url(source_url):
            stats["scanned"] += 1
            crawler_logger.info(
                "run_id=%s stage=direct_detail_mode source=%s detail_url=%s",
                run_id,
                source_name,
                source_url,
            )
            try:
                detail_html = fetch_html_with_retries(source_url, run_id, db=db, source=source_name)
                detail = parse_notice_detail(detail_html)
            except Exception as exc:
                stats["fetchErrors"] += 1
                crawler_logger.warning(
                    "run_id=%s stage=direct_detail_fetch_failed source=%s url=%s error=%s",
                    run_id,
                    source_name,
                    source_url,
                    exc,
                )
                continue

            if db:
                _record_source_success(db, source_name)

            title_hint = _title_hint_from_url(source_url)
            query_hint = _query_hint_from_url(source_url)
            detail_title = detail.get("title") or title_hint or source_url
            detail_text = detail.get("body") or title_hint or ""
            publish_date = detail.get("publishDate") or ""
            company = detail.get("company") or ""
            if company and not _is_valid_company(company):
                company = ""
            region = detail.get("region") or ""
            contact_name = _extract_contact_name(detail_text)
            contact_phone = _extract_contact_phone(detail_text)

            combined_text = f"{detail_title} {detail_text} {title_hint} {query_hint}"
            matched, reason = _matches_crawler_rules(combined_text, keywords, exclude_keywords)
            if not matched:
                stats["filtered"] += 1
                crawler_logger.info(
                    "run_id=%s stage=direct_detail_filter source=%s action=filtered reason=%s url=%s",
                    run_id,
                    source_name,
                    reason,
                    source_url,
                )
                continue

            budget_wan = _extract_budget(combined_text)
            dimension_matched, dimension_reason = _matches_dimension_filters(
                combined_text,
                region,
                budget_wan,
                publish_date,
                search_rules,
            )
            if not dimension_matched:
                stats["filtered"] += 1
                crawler_logger.info(
                    "run_id=%s stage=direct_detail_filter source=%s action=filtered reason=%s url=%s",
                    run_id,
                    source_name,
                    dimension_reason,
                    source_url,
                )
                continue
            scores = _score_from_notice(combined_text, budget_wan)
            requirements = (detail_text or title_hint)[:1800]
            if budget_wan is not None:
                requirements = f"预算金额约 {budget_wan} 万元。\n{requirements}"
            # 需求详情不再附带原文链接

            candidates.append({
                "title": detail_title,
                "company": company or "待核验采购单位",
                "publishDate": publish_date,
                "region": region or "待核验地区",
                "industry": _pick(industries, len(candidates), DEFAULT_CRAWLER_INDUSTRIES),
                "category": _pick(categories, len(candidates), DEFAULT_CRAWLER_CATEGORIES),
                "sourceDb": source_name,
                "detailUrl": source_url,
                "requirements": requirements,
                "contactName": contact_name or None,
                "contactPhone": contact_phone or None,
                **scores,
            })
            crawler_logger.info(
                "run_id=%s stage=direct_detail_accept source=%s title=%s url=%s",
                run_id,
                source_name,
                detail_title,
                source_url,
            )
            continue

        try:
            notices, pages_ok, has_links = fetch_paginated_notice_links(
                source_name, source_url, run_id, db=db, search_rules=search_rules
            )
            crawler_logger.info(
                "run_id=%s stage=source_list_paginated source=%s pages_ok=%s links=%s",
                run_id,
                source_name,
                pages_ok,
                len(notices),
            )
        except Exception as exc:
            stats["fetchErrors"] += 1
            crawler_logger.warning(
                "run_id=%s stage=source_fetch_failed source=%s url=%s error=%s",
                run_id,
                source_name,
                source_url,
                exc,
            )
            continue

        if db:
            _record_source_success(db, source_name, structure_valid=has_links)

        for index, notice in enumerate(notices, start=1):
            stats["scanned"] += 1
            detail_text = notice["listSummary"]
            detail_title = notice["title"]
            publish_date = notice.get("publishDate") or ""
            company = ""
            region = ""
            contact_name = ""
            contact_phone = ""

            try:
                crawler_logger.info(
                    "run_id=%s stage=detail_fetch_start source=%s index=%s detail_url=%s",
                    run_id,
                    source_name,
                    index,
                    notice["detailUrl"],
                )
                detail_html = fetch_html(notice["detailUrl"], run_id)
                detail = parse_notice_detail(detail_html)
                detail_title = detail["title"] or detail_title
                detail_text = detail["body"] or detail_text
                publish_date = detail["publishDate"] or publish_date
                company = detail["company"]
                if company and not _is_valid_company(company):
                    company = ""
                region = detail["region"]
                contact_name = _extract_contact_name(detail_text)
                contact_phone = _extract_contact_phone(detail_text)
            except Exception as exc:
                stats["fetchErrors"] += 1
                crawler_logger.info(
                    "run_id=%s stage=detail_fetch_failed source=%s index=%s url=%s error=%s",
                    run_id,
                    source_name,
                    index,
                    notice["detailUrl"],
                    exc,
                )

            query_hint = _query_hint_from_url(notice["detailUrl"])
            combined_text = f"{detail_title} {detail_text} {query_hint}"
            matched, reason = _matches_crawler_rules(combined_text, keywords, exclude_keywords)
            if not matched:
                stats["filtered"] += 1
                crawler_logger.info(
                    "run_id=%s stage=web_filter source=%s index=%s action=filtered reason=%s title=%s",
                    run_id,
                    source_name,
                    index,
                    reason,
                    detail_title,
                )
                continue

            budget_wan = _extract_budget(combined_text)
            dimension_matched, dimension_reason = _matches_dimension_filters(
                combined_text,
                region,
                budget_wan,
                publish_date,
                search_rules,
            )
            if not dimension_matched:
                stats["filtered"] += 1
                crawler_logger.info(
                    "run_id=%s stage=web_filter source=%s index=%s action=filtered reason=%s title=%s",
                    run_id,
                    source_name,
                    index,
                    dimension_reason,
                    detail_title,
                )
                continue
            scores = _score_from_notice(combined_text, budget_wan)
            requirements = detail_text[:1800]
            if budget_wan is not None:
                requirements = f"预算金额约 {budget_wan} 万元。\n{requirements}"
            # 需求详情不再附带原文链接

            candidates.append({
                "title": detail_title,
                "company": company or "待核验采购单位",
                "publishDate": publish_date,
                "region": region or "待核验地区",
                "industry": _pick(industries, len(candidates), DEFAULT_CRAWLER_INDUSTRIES),
                "category": _pick(categories, len(candidates), DEFAULT_CRAWLER_CATEGORIES),
                "sourceDb": source_name,
                "detailUrl": notice["detailUrl"],
                "requirements": requirements,
                "contactName": contact_name or None,
                "contactPhone": contact_phone or None,
                **scores,
            })
            crawler_logger.info(
                "run_id=%s stage=web_accept source=%s index=%s title=%s detail_url=%s",
                run_id,
                source_name,
                index,
                detail_title,
                notice["detailUrl"],
            )

    return candidates, stats

def build_crawler_candidates(
    search_rules: Dict[str, Any],
    run_id: str,
    db: Optional[Session] = None,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    return collect_web_candidates(search_rules, run_id, db=db)

def attach_scores_and_stage_leads(
    db: Session,
    candidates: List[Dict[str, Any]],
    rules: DBScoringRule,
    run_id: str,
) -> Tuple[List[Dict[str, Any]], int]:
    existing_keys = {
        _lead_identity_key(title or "", company or "")
        for title, company in db.query(DBLead.title, DBLead.company).yield_per(1000)
    }
    integrated_leads: List[Dict[str, Any]] = []
    duplicate_count = 0

    for index, candidate in enumerate(candidates, start=1):
        identity_key = _lead_identity_key(candidate.get("title", ""), candidate.get("company", ""))
        if identity_key in existing_keys:
            # 重复线索：更新 source_db 为 detailUrl（如果有）
            detail_url = candidate.get("detailUrl")
            if detail_url:
                existing = db.query(DBLead).filter(
                    DBLead.title == candidate.get("title", ""),
                    DBLead.company == candidate.get("company", ""),
                ).first()
                if existing and not (existing.source_db or "").startswith("http"):
                    existing.source_db = _clamp_str(detail_url, 150)
            duplicate_count += 1
            crawler_logger.info(
                "run_id=%s stage=dedupe candidate=%s action=skipped_duplicate title=%s",
                run_id,
                index,
                candidate.get("title", ""),
            )
            continue

        budget_score = _coerce_int(candidate.get("budgetScore"), 50)
        match_score = _coerce_int(candidate.get("matchScore"), 50)
        qualification_score = _coerce_int(candidate.get("qualificationScore"), 50)
        stage_score = _coerce_int(candidate.get("stageScore"), 50)
        region_score = _coerce_int(candidate.get("regionScore"), 50)
        total_score, grade = calculate_score(
            rules,
            budget_score,
            match_score,
            stage_score,
            qualification_score,
            region_score,
        )

        db_lead = DBLead(
            id=str(uuid.uuid4()),
            title=_clamp_str(candidate.get("title", ""), 255),
            company=_clamp_str(candidate.get("company", ""), 255) or "待核验采购单位",
            publish_date=date.fromisoformat(candidate["publishDate"]) if candidate.get("publishDate") else date.today(),
            region=_sanitize_region(candidate.get("region", "")) or "待核验地区",
            industry=_clamp_str(candidate.get("industry", ""), 100),
            category=_clamp_str(candidate.get("category", ""), 100),
            source_db=_clamp_str(candidate.get("detailUrl") or candidate.get("sourceDb", "公网抓取"), 150),
            requirements=candidate.get("requirements", ""),
            contact_name=_clamp_str(candidate.get("contactName"), 100) or None,
            contact_phone=_clamp_str(candidate.get("contactPhone"), 30) or None,
            budget_score=budget_score,
            match_score=match_score,
            qualification_score=qualification_score,
            stage_score=stage_score,
            region_score=region_score,
            total_score=total_score,
            grade=grade,
            status="待分级",
            crawl_time=_now(),
            lead_source="crawler",
            created_at=_now(),
        )
        db_lead.status = "待分配"
        db.add(db_lead)
        existing_keys.add(identity_key)
        integrated_leads.append(lead_to_api(db_lead))
        crawler_logger.info(
            "run_id=%s stage=scoring candidate=%s total_score=%.2f grade=%s",
            run_id,
            index,
            total_score,
            grade,
        )

    return integrated_leads, duplicate_count

# --- 审计日志 ---
@app.get("/api/audit-logs")
def get_audit_logs(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role in ("超级管理员", "商务负责人"):
        query = db.query(DBAuditLog)
    else:
        query = db.query(DBAuditLog).filter(DBAuditLog.operator == current_user["name"])
    total = query.count()
    logs = query.order_by(DBAuditLog.timestamp.desc()).limit(200).all()
    return {"total": total, "items": [audit_to_api(l) for l in logs]}

# --- 爬虫/抓取 API 路由（含完整日志，PRD 4.1 和用户需求） ---
@app.post("/api/gemini/scrape")
def api_gemini_scrape(payload: Optional[dict] = None, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权执行网监抓取")

    run_id = str(uuid.uuid4())[:8]
    started_at = time.perf_counter()
    crawler_logger.info("run_id=%s stage=start timestamp=%s", run_id, _now().isoformat())

    try:
        db_config = search_config_to_api(db.query(DBSearchConfig).first())
        override_config = payload.get("searchConfig") if isinstance(payload, dict) else None
        search_rules = normalize_search_config(override_config or db_config)

        candidates, scan_stats = build_crawler_candidates(search_rules, run_id, db=db)
        rules = get_or_create_scoring_rule(db)
        integrated_leads, duplicate_count = attach_scores_and_stage_leads(db, candidates, rules, run_id)

        crawler_logger.info(
            "run_id=%s stage=persistence action=commit created=%s filtered=%s duplicates=%s",
            run_id,
            len(integrated_leads),
            scan_stats["filtered"],
            duplicate_count,
        )
        db.commit()
        duration_ms = int((time.perf_counter() - started_at) * 1000)

        try:
            add_audit_log(
                db,
                "系统AI引擎",
                "系统管理员",
                "规则引擎抓取",
                (
                    f"[爬虫接口监测日志] run_id={run_id}，基于核心词库 "
                    f"[{'/'.join(search_rules['keywords'] or DEFAULT_CRAWLER_KEYWORDS)}] 与黑名单 "
                    f"[{'/'.join(search_rules['excludeKeywords']) or '无'}] 完成扫描。"
                    f"扫描 {scan_stats['scanned']} 条，过滤 {scan_stats['filtered']} 条，"
                    f"重复跳过 {duplicate_count} 条，新增入库 {len(integrated_leads)} 条，耗时 {duration_ms}ms。"
                ),
            )
        except Exception:
            db.rollback()
            crawler_logger.exception("run_id=%s stage=audit_log_failed_after_commit", run_id)

        crawler_logger.info(
            "run_id=%s stage=success duration_ms=%s created=%s",
            run_id,
            duration_ms,
            len(integrated_leads),
        )
        return {
            "success": True,
            "apiUsed": False,
            "runId": run_id,
            "data": integrated_leads,
            "stats": {
                "mode": scan_stats.get("mode", "web"),
                "scannedCount": scan_stats["scanned"],
                "filteredCount": scan_stats["filtered"],
                "skippedDuplicateCount": duplicate_count,
                "createdCount": len(integrated_leads),
                "sourceCount": scan_stats.get("sourceCount", 0),
                "fetchErrors": scan_stats.get("fetchErrors", 0),
                "unresolvedSources": scan_stats.get("unresolvedSources", []),
                "durationMs": duration_ms,
            },
        }
    except Exception as exc:
        db.rollback()
        duration_ms = int((time.perf_counter() - started_at) * 1000)
        crawler_logger.exception(
            "run_id=%s stage=failed duration_ms=%s error=%s",
            run_id,
            duration_ms,
            exc,
        )
        try:
            add_audit_log(
                db,
                "系统AI引擎",
                "系统管理员",
                "规则引擎抓取失败",
                f"[爬虫接口错误日志] run_id={run_id}，执行失败，耗时 {duration_ms}ms，错误：{exc}",
            )
        except Exception:
            db.rollback()
            crawler_logger.exception("run_id=%s stage=audit_log_failed", run_id)
        raise HTTPException(status_code=500, detail="Crawler execution failed") from exc

# --- 竞品管理 ---
def _normalize_competitor_name(name: str) -> str:
    return "".join((name or "").split()).lower()


@app.get("/api/competitors")
def get_competitors(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    competitors = db.query(DBCompetitor).all()
    return [competitor_to_api(c) for c in competitors]


@app.post("/api/competitors")
def create_competitor(comp: CompetitorCreate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权新增竞品")

    name = (comp.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="竞品名称不能为空")

    # 重复检测：同名（去空格、大小写不敏感）且仍有效的竞品视为重复，返回 409 供前端合并。
    target_key = _normalize_competitor_name(name)
    for existing in db.query(DBCompetitor).all():
        if (getattr(existing, "status", "active") or "active") != "active":
            continue
        if _normalize_competitor_name(existing.name) == target_key:
            raise HTTPException(
                status_code=409,
                detail={
                    "message": f"已存在同名竞品「{existing.name}」，建议合并或改为编辑该竞品。",
                    "existingId": existing.id,
                    "existingName": existing.name,
                },
            )

    db_comp = DBCompetitor(
        id=str(uuid.uuid4()),
        name=name,
        strengths=json.dumps(comp.strengths, ensure_ascii=False),
        weaknesses=json.dumps(comp.weaknesses, ensure_ascii=False),
        pricing_range=comp.pricingRange,
        win_rate=comp.winRate,
        main_business=comp.mainBusiness,
        tech_route=comp.techRoute,
        typical_cases=comp.typicalCases,
        info_source=comp.infoSource,
        priority=comp.priority or "normal",
        updated_at=_now(),
        status="active",
    )
    db.add(db_comp)
    db.commit()
    db.refresh(db_comp)
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "新增竞品",
        f"新增竞品「{db_comp.name}」",
        db_comp.id,
    )
    return competitor_to_api(db_comp)


@app.put("/api/competitors/{comp_id}")
def update_competitor(comp_id: str, patch: CompetitorUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权修改竞品")

    db_comp = db.query(DBCompetitor).filter(DBCompetitor.id == comp_id).first()
    if not db_comp:
        raise HTTPException(status_code=404, detail="竞品不存在")

    if patch.name is not None and patch.name.strip():
        db_comp.name = patch.name.strip()
    if patch.strengths is not None:
        db_comp.strengths = json.dumps(patch.strengths, ensure_ascii=False)
    if patch.weaknesses is not None:
        db_comp.weaknesses = json.dumps(patch.weaknesses, ensure_ascii=False)
    if patch.pricingRange is not None:
        db_comp.pricing_range = patch.pricingRange
    if patch.winRate is not None:
        db_comp.win_rate = patch.winRate
    if patch.mainBusiness is not None:
        db_comp.main_business = patch.mainBusiness
    if patch.techRoute is not None:
        db_comp.tech_route = patch.techRoute
    if patch.typicalCases is not None:
        db_comp.typical_cases = patch.typicalCases
    if patch.infoSource is not None:
        db_comp.info_source = patch.infoSource
    if patch.status is not None and patch.status in ("active", "inactive"):
        db_comp.status = patch.status
    if patch.priority is not None and patch.priority in ("key", "normal"):
        db_comp.priority = patch.priority
    db_comp.updated_at = _now()

    db.commit()
    db.refresh(db_comp)
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "更新竞品",
        f"更新竞品「{db_comp.name}」资料",
        db_comp.id,
    )
    return competitor_to_api(db_comp)


@app.put("/api/competitors/{comp_id}/status")
def update_competitor_status(comp_id: str, patch: CompetitorStatusUpdate, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权修改竞品状态")

    db_comp = db.query(DBCompetitor).filter(DBCompetitor.id == comp_id).first()
    if not db_comp:
        raise HTTPException(status_code=404, detail="竞品不存在")
    new_status = patch.status if patch.status in ("active", "inactive") else "active"
    db_comp.status = new_status
    db_comp.updated_at = _now()
    db.commit()
    db.refresh(db_comp)
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "竞品失效" if new_status == "inactive" else "竞品启用",
        f"将竞品「{db_comp.name}」标记为{'失效' if new_status == 'inactive' else '有效'}",
        db_comp.id,
    )
    return competitor_to_api(db_comp)

@app.delete("/api/competitors/{comp_id}")
def delete_competitor(comp_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权删除竞品")

    db_comp = db.query(DBCompetitor).filter(DBCompetitor.id == comp_id).first()
    if not db_comp:
        raise HTTPException(status_code=404, detail="竞品不存在")

    # 级联删除竞品更新留痕
    db.query(DBCompetitorUpdate).filter(DBCompetitorUpdate.competitor_id == comp_id).delete()
    db.delete(db_comp)
    db.commit()
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "删除竞品",
        f"删除竞品「{db_comp.name}」及其更新留痕",
        comp_id,
    )
    return {"ok": True}

# --- 模板管理 ---
@app.get("/api/templates")
def get_templates(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    templates = db.query(DBTemplate).all()
    return [template_to_api(t) for t in templates]

@app.post("/api/clean-duplicates")
def clean_duplicates(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人"):
        raise HTTPException(status_code=403, detail="仅超级管理员或商务负责人有权执行自洁去重")
        
    leads = db.query(DBLead).order_by(DBLead.publish_date.desc()).all()
    seen = set()
    removed = 0
    for lead in leads:
        key = f"{(lead.title or '').strip()}::{(lead.company or '').strip()}"
        if not key.strip(" :"):
            continue
        if key in seen:
            db.delete(lead)
            removed += 1
        else:
            seen.add(key)
    db.commit()
    add_audit_log(db, current_user["name"], current_user["role"], "自洁去重", f"执行线索去重，移除 {removed} 条重复记录。")
    return {"removedCount": removed}

# ============ 预调研：多源聚合 + 痛点报告 + 版本化归档 ============


def _extract_original_link(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m = re.search(r"原文链接[：:]\s*(https?://[^\s]+)", text)
    if m:
        return m.group(1).strip()
    m = re.search(r"(https?://[^\s]+)", text)
    return m.group(1).strip() if m else None


def _find_keyword_context(text: str, keywords: List[str], window: int = 40) -> Optional[str]:
    if not text:
        return None
    for kw in keywords:
        idx = text.find(kw)
        if idx >= 0:
            start = max(0, idx - window)
            end = min(len(text), idx + len(kw) + window)
            snippet = text[start:end].strip()
            return re.sub(r"\s+", " ", snippet)
    return None


def aggregate_public_info(db: Session, lead: DBLead) -> Dict[str, Any]:
    """复用系统已抓取的招投标线索数据，聚合目标客户/项目的结构化原始信息集。

    所有信息标注来源链接；公开信息不足时如实标记缺失，禁止编造。
    如启用 DeepSeek 联网搜索，LLM 生成报告时会自行补充公网信息。
    """
    company = (lead.company or "").strip()
    
    # 如果采购单位为"待核验采购单位"，尝试从 requirements 文本中重新提取
    if company in ("", "待核验采购单位", "0"):
        re_extracted = _extract_field(lead.requirements or "", [
            "项目法人（招标人）", "项目法人(招标人)", "采购人名称", "招标人名称", "采购单位名称",
            "采购人（甲方）", "采购人(甲方)",
            "采购人", "招标人", "采购单位", "建设单位", "项目业主", "甲方",
        ])
        if re_extracted:
            company = re_extracted
    
    self_link = _extract_original_link(lead.requirements) or _extract_original_link(lead.notes)

    related: List[DBLead] = []
    if company:
        related = (
            db.query(DBLead)
            .filter(DBLead.company == company)
            .order_by(DBLead.publish_date.desc())
            .all()
        )
    if all(r.id != lead.id for r in related):
        related = [lead] + related

    combined_text = " ".join(
        [r.requirements or "" for r in related] + [r.title or "" for r in related]
    )

    field_label = dict(RESEARCH_INFO_FIELDS)
    info: Dict[str, Any] = {}
    missing: List[str] = []
    sources: List[Dict[str, str]] = []

    if company:
        info["companyBasic"] = {
            "value": f"{company}（地区：{lead.region or '未知'}）",
            "source": self_link,
        }
    else:
        missing.append(field_label["companyBasic"])

    field_keywords = {
        "plantScale": ["规模", "万吨", "处理能力", "吨/日", "吨/天", "m³/d", "立方"],
        "treatmentProcess": ["工艺", "AAO", "A2O", "AO", "MBR", "氧化沟", "SBR", "曝气", "生化"],
        "pastProjects": ["改造", "提升", "扩建", "新建", "EPC", "总承包"],
        "operationMode": ["运维", "运营", "托管", "委托运营", "BOT", "PPP", "特许经营"],
        "energyLaborCost": ["能耗", "电耗", "吨水电耗", "人工", "人员", "成本", "费用"],
    }
    for key, kws in field_keywords.items():
        ctx = _find_keyword_context(combined_text, kws)
        if ctx:
            info[key] = {"value": ctx, "source": self_link}
        else:
            missing.append(field_label[key])

    history = []
    for r in related:
        link = _extract_original_link(r.requirements)
        history.append({
            "title": r.title,
            "publishDate": _date_str(r.publish_date) or "",
            "category": r.category or "",
            "sourceDb": r.source_db or "",
            "link": link,
        })
        if link:
            short_title = r.title[:40] + "…" if len(r.title) > 40 else r.title
            sources.append({"label": f"招投标公告·{short_title}", "url": link})
    info["history"] = history

    seen = set()
    uniq_sources = []
    for s in sources:
        if s["url"] and s["url"] not in seen:
            seen.add(s["url"])
            uniq_sources.append(s)

    total_fields = len(RESEARCH_INFO_FIELDS)
    missing_field_count = len([m for m in missing if m in field_label.values()])
    completeness = int(round((total_fields - missing_field_count) / total_fields * 100)) if total_fields else 0

    return {
        "company": company,
        "info": info,
        "missing": missing,
        "sources": uniq_sources,
        "relatedCount": len(related),
        "completeness": completeness,
        "combinedText": combined_text,
    }


def _match_capabilities(text: str) -> List[Dict[str, str]]:
    lowered = (text or "").lower()
    matched = [c for c in BUSINESS_CAPABILITIES if any(kw.lower() in lowered for kw in c["keywords"])]
    if not matched:
        matched = [c for c in BUSINESS_CAPABILITIES if c["name"] == "综合智能化改造"]
    return matched


def build_pain_point_report(lead: DBLead, agg: Dict[str, Any], guide: str = "") -> Dict[str, Any]:
    company = agg["company"] or (lead.company or "目标客户")
    info = agg["info"]
    missing = agg["missing"]
    completeness = agg["completeness"]
    combined_text = agg["combinedText"]

    sufficient = bool(company) and (completeness >= 34 or agg["relatedCount"] >= 1)
    matched_caps = _match_capabilities(combined_text)

    def line(label: str, key: str) -> str:
        v = info.get(key)
        return f"- {label}：{v['value']}" if v else f"- {label}：【缺失，建议补充调研】"

    sections = []
    overview = [
        f"- 目标客户：{company}",
        f"- 所在地区：{lead.region or '未知'}",
        f"- 当前项目：{lead.title or '未知'}",
        line("厂区规模", "plantScale"),
        line("处理工艺", "treatmentProcess"),
        f"- 历史相关公告：{agg['relatedCount']} 条",
    ]
    sections.append(("一、企业及项目概况", "\n".join(overview)))

    pains = []
    if info.get("energyLaborCost"):
        pains.append(f"- 公开信息信号：{info['energyLaborCost']['value']}")
    if info.get("operationMode"):
        pains.append(f"- 运维模式信号：{info['operationMode']['value']}")
    pains.append("- 污水行业共性痛点：")
    pains.extend([f"    · {p}" for p in WATER_INDUSTRY_PAIN_POINTS])
    sections.append(("二、现存运维/能耗/成本痛点分析", "\n".join(pains)))

    entry = [f"- 【{c['name']}】{c['value']}" for c in matched_caps]
    sections.append(("三、我方业务适配切入点", "\n".join(entry)))

    advice = [
        f"- 建议优先切入：{matched_caps[0]['name']}，以「可量化降本」作为首轮沟通钩子。",
        "- 首轮沟通建议携带同类项目改造收益数据与现场调研清单。",
        "- 关注预算节点与招标节奏，提前对接设计院 / 代理机构关键人。",
    ]
    if guide:
        advice.insert(0, f"- 本次调研倾向约束：{guide}")
    sections.append(("四、对接沟通建议", "\n".join(advice)))

    todo = [f"- {m}" for m in missing] or ["- 暂无明显缺失项，可进入现场深度调研。"]
    sections.append(("五、需进一步确认的信息清单", "\n".join(todo)))

    body = "\n\n".join([f"### {t}\n{c}" for t, c in sections])
    header = (
        f"# {company} 客户预调研报告\n"
        f"> 信息完整度：{completeness}%　|　数据来源：系统已抓取招投标线索（共 {agg['relatedCount']} 条）"
    )
    if not sufficient:
        body = "> ⚠️ 公开信息不足，建议补充调研。以下仅为基础信息汇总：\n\n" + body
    return {
        "content": header + "\n\n" + body,
        "sufficient": sufficient,
        "completeness": completeness,
    }


def _maybe_llm_report(company: str, lead: DBLead, agg: Dict[str, Any], guide: str) -> Optional[str]:
    """配置了 LLM_API_KEY 时调用大模型生成；否则返回 None 走规则引擎。"""
    if not LLM_API_KEY:
        logger.info("[AI引擎] 未检测到 LLM_API_KEY 环境变量，跳过 AI 模型调用，使用规则引擎生成预调研报告")
        return None

    logger.info(
        "[AI引擎] === 开始调用 AI 模型生成预调研 ==="
        "\n  ├─ 目标客户: %s"
        "\n  ├─ 当前项目: %s"
        "\n  ├─ 模型配置: %s @ %s"
        "\n  ├─ 信息完整度: %s%%"
        "\n  └─ 调研约束: %s",
        company, lead.title, LLM_MODEL, LLM_API_BASE, agg["completeness"], guide or "无"
    )

    try:
        caps_desc = "\n".join([f"- {c['name']}：{c['value']}" for c in BUSINESS_CAPABILITIES])
        info_desc = json.dumps(agg["info"], ensure_ascii=False, indent=2)
        sources_desc = json.dumps(agg["sources"], ensure_ascii=False)
        source_hint = "（暂无补充数据，可联网搜索）" if ENABLE_DEEPSEEK_SEARCH else "（仅系统内部招投标数据）"
        user_prompt = (
            f"目标客户：{company}\n所在地区：{lead.region or '未知'}\n当前项目：{lead.title or '未知'}\n"
            f"信息完整度：{agg['completeness']}%\n缺失项：{agg['missing']}\n\n"
            f"【已聚合的公开信息{source_hint}（含来源，禁止编造未提供的信息）】\n{info_desc}\n\n"
            f"【来源链接】\n{sources_desc}\n\n"
            f"【我方业务能力库】\n{caps_desc}\n\n"
            f"调研倾向约束：{guide or '无'}\n\n"
            "请严格按以下固定结构输出 Markdown 报告：一、企业及项目概况；二、现存运维/能耗/成本痛点分析（结合污水行业共性）；"
            "三、我方业务适配切入点（匹配上述业务能力库）；四、对接沟通建议；五、需进一步确认的信息清单。"
            "缺失信息必须如实列在第五部分，禁止编造。"
        )

        request_body = {
            "model": LLM_MODEL,
            "messages": [
                {"role": "system", "content": "你是污水处理行业资深售前顾问，只能基于用户提供的公开信息撰写客户预调研报告，禁止编造。"},
                {"role": "user", "content": user_prompt},
            ],
            "temperature": 0.4,
        }

        if ENABLE_DEEPSEEK_SEARCH:
            request_body["search"] = True
            logger.info("[AI引擎] 已启用 DeepSeek 联网搜索，AI 将自动搜索补全企业官网/行业新闻/环保公示等公开信息")

        resp = requests.post(
            f"{LLM_API_BASE}/chat/completions",
            headers={"Authorization": f"Bearer {LLM_API_KEY}", "Content-Type": "application/json"},
            json=request_body,
            timeout=LLM_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        content = (data.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()

        if content:
            total_tokens = (data.get("usage") or {}).get("total_tokens", "未知")
            logger.info(
                "[AI引擎] ✅ AI 模型调用成功"
                "\n  ├─ 模型: %s"
                "\n  ├─ 消耗 tokens: %s"
                "\n  ├─ 报告长度: %s 字"
                "\n  └─ 结果: 使用 LLM 生成的预调研报告",
                LLM_MODEL, total_tokens, len(content)
            )
            return content

        logger.warning("[AI引擎] ⚠️ AI 模型返回内容为空，回退规则引擎")
        return None
    except Exception as exc:
        logger.warning("[AI引擎] ❌ AI 模型调用失败（回退规则引擎）：%s", exc)
        return None


def research_to_api(r: DBResearchReport) -> dict:
    def _loads(raw, default):
        try:
            return json.loads(raw) if raw else default
        except Exception:
            return default
    return {
        "id": r.id,
        "leadId": r.lead_id,
        "version": r.version,
        "createdAt": _dt_str(r.created_at),
        "operator": r.operator,
        "engine": r.engine,
        "completeness": int(r.completeness or 0),
        "aggregatedInfo": _loads(r.aggregated_info, {}),
        "sources": _loads(r.sources, []),
        "missingItems": _loads(r.missing_items, []),
        "reportContent": r.report_content or "",
        "isSufficient": bool(r.is_sufficient),
    }


def run_pre_research(db: Session, lead: DBLead, guide: str, operator: str) -> DBResearchReport:
    agg = aggregate_public_info(db, lead)
    company = agg["company"] or (lead.company or "目标客户")

    llm_text = _maybe_llm_report(company, lead, agg, guide)
    if llm_text:
        report_text = llm_text
        engine = "llm"
        completeness = agg["completeness"]
        sufficient = bool(company) and (completeness >= 34 or agg["relatedCount"] >= 1)
        logger.info("[AI引擎] 引擎决策结果 → 使用 AI 模型（LLM）")
    else:
        built = build_pain_point_report(lead, agg, guide)
        report_text = built["content"]
        engine = "rule"
        completeness = built["completeness"]
        sufficient = built["sufficient"]
        logger.info("[AI引擎] 引擎决策结果 → 使用规则引擎（Rule）")

    last = (
        db.query(DBResearchReport)
        .filter(DBResearchReport.lead_id == lead.id)
        .order_by(DBResearchReport.version.desc())
        .first()
    )
    version = (last.version + 1) if last else 1

    record = DBResearchReport(
        id=str(uuid.uuid4()),
        lead_id=lead.id,
        version=version,
        operator=operator,
        engine=engine,
        completeness=completeness,
        aggregated_info=json.dumps(agg["info"], ensure_ascii=False),
        sources=json.dumps(agg["sources"], ensure_ascii=False),
        missing_items=json.dumps(agg["missing"], ensure_ascii=False),
        report_content=report_text,
        is_sufficient=1 if sufficient else 0,
    )
    db.add(record)
    lead.pre_research_report = report_text
    lead.research_date = date.today()
    db.commit()
    db.refresh(record)
    add_audit_log(
        db,
        "系统AI引擎",
        operator,
        "生成预调研报告",
        f"为线索「{lead.title}」生成第 {version} 版预调研报告（引擎：{'大模型' if engine == 'llm' else '规则引擎'}，完整度 {completeness}%）",
        lead.id,
    )
    return record


@app.post("/api/leads/{lead_id}/research")
def create_research(lead_id: str, payload: dict = None, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人", "商务专员"):
        raise HTTPException(status_code=403, detail="仅超级管理员、商务负责人或商务专员有权执行预调研")

    lead = db.query(DBLead).filter(DBLead.id == lead_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="线索不存在")
    payload = payload or {}
    guide = (payload.get("customGuide") or payload.get("guide") or "").strip()
    record = run_pre_research(db, lead, guide, current_user["name"])
    return research_to_api(record)


@app.get("/api/leads/{lead_id}/research")
def list_research(lead_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    rows = (
        db.query(DBResearchReport)
        .filter(DBResearchReport.lead_id == lead_id)
        .order_by(DBResearchReport.version.desc())
        .all()
    )
    return [research_to_api(r) for r in rows]


# ============ 竞品对标分析：智能补全 + 项目级对标报告（混合引擎） ============

def _extract_json_block(text: str) -> Optional[dict]:
    """从大模型返回文本中尽量提取 JSON 对象，做好容错。"""
    if not text:
        return None
    # 优先剥离 ```json ... ``` 代码块
    fenced = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", text)
    candidate = fenced.group(1) if fenced else None
    if candidate is None:
        start = text.find("{")
        end = text.rfind("}")
        if start >= 0 and end > start:
            candidate = text[start:end + 1]
    if candidate is None:
        return None
    try:
        parsed = json.loads(candidate)
        return parsed if isinstance(parsed, dict) else None
    except Exception:
        return None


COMPETITOR_INTEL_CATEGORIES = [
    ("newProducts", "新品发布"),
    ("winningBids", "中标公告"),
    ("techNews", "技术动态"),
    ("pricingChanges", "报价调整"),
    ("customerCases", "客户案例"),
]


def _normalize_intel_items(items: Any) -> List[Dict[str, str]]:
    """将 LLM 返回的单项/列表规范为 [{title, summary, source, date}, ...]。"""
    if not items:
        return []
    if isinstance(items, str) and items.strip():
        return [{"title": "", "summary": items.strip(), "source": "", "date": ""}]
    if not isinstance(items, list):
        return []
    normalized: List[Dict[str, str]] = []
    for item in items:
        if isinstance(item, str) and item.strip():
            normalized.append({"title": "", "summary": item.strip(), "source": "", "date": ""})
        elif isinstance(item, dict):
            normalized.append({
                "title": str(item.get("title") or "").strip(),
                "summary": str(item.get("summary") or item.get("content") or "").strip(),
                "source": str(item.get("source") or item.get("url") or "").strip(),
                "date": str(item.get("date") or item.get("time") or "").strip(),
            })
    return [x for x in normalized if x.get("title") or x.get("summary")]


def _normalize_intel_payload(raw: Optional[dict]) -> Dict[str, Any]:
    payload = raw or {}
    intel: Dict[str, Any] = {}
    for key, _label in COMPETITOR_INTEL_CATEGORIES:
        intel[key] = _normalize_intel_items(payload.get(key))
    intel["infoSource"] = str(payload.get("infoSource") or "").strip()
    return intel


def _intel_has_content(intel: Dict[str, Any]) -> bool:
    for key, _ in COMPETITOR_INTEL_CATEGORIES:
        if intel.get(key):
            return True
    return bool(intel.get("infoSource"))


def _enrich_competitor_via_llm(comp: DBCompetitor) -> Optional[dict]:
    """调用 DeepSeek 全网检索竞品动态；未配置 key 或失败时返回 None。"""
    if not LLM_API_KEY:
        logger.info("[竞品AI] 未检测到 LLM_API_KEY，跳过竞品 AI 智能补全")
        return None

    logger.info(
        "[竞品AI] === 开始调用 AI 检索竞品动态 ==="
        "\n  ├─ 竞品名称: %s"
        "\n  ├─ 模型配置: %s"
        "\n  └─ 联网搜索: %s",
        comp.name, LLM_MODEL, "已开启" if ENABLE_DEEPSEEK_SEARCH else "未开启"
    )

    try:
        known = {
            "name": comp.name,
            "mainBusiness": comp.main_business or "",
            "techRoute": comp.tech_route or "",
            "pricingRange": comp.pricing_range or "",
            "typicalCases": comp.typical_cases or "",
            "strengths": _safe_json_list(comp.strengths),
            "weaknesses": _safe_json_list(comp.weaknesses),
            "infoSource": comp.info_source or "",
            "existingIntel": _safe_json_dict(getattr(comp, "monitoring_intel", None)),
        }
        user_prompt = (
            f"竞品名称：{comp.name}\n"
            f"行业领域：污水智能化改造（智能加药、智能曝气、少人化运维、综合改造等）\n"
            f"【已知资料（仅供参考，请检索补充最新公开动态，禁止编造）】\n"
            f"{json.dumps(known, ensure_ascii=False, indent=2)}\n\n"
            "请基于公开可得信息，检索并整理该竞品近期动态，严格只返回如下 JSON（不要 Markdown、不要解释）：\n"
            "{\n"
            '  "newProducts": [{"title":"", "summary":"", "source":"来源链接或出处", "date":"YYYY-MM-DD或空"}],\n'
            '  "winningBids": [{"title":"", "summary":"", "source":"", "date":""}],\n'
            '  "techNews": [{"title":"", "summary":"", "source":"", "date":""}],\n'
            '  "pricingChanges": [{"title":"", "summary":"", "source":"", "date":""}],\n'
            '  "customerCases": [{"title":"", "summary":"", "source":"", "date":""}],\n'
            '  "infoSource": "本次检索总体来源说明",\n'
            '  "mainBusiness": "主营业务(如有新信息)",\n'
            '  "techRoute": "技术路线(如有新信息)",\n'
            '  "pricingRange": "报价水平(如有新信息)",\n'
            '  "typicalCases": "典型案例(如有新信息)",\n'
            '  "strengths": ["优势"],\n'
            '  "weaknesses": ["劣势"]\n'
            "}\n"
            "五类动态若无公开信息则返回空数组 []；无法确认的基础字段返回空字符串或空数组，禁止编造。"
        )
        resp = requests.post(
            f"{LLM_API_BASE}/chat/completions",
            headers={"Authorization": f"Bearer {LLM_API_KEY}", "Content-Type": "application/json"},
            json={
                "model": LLM_MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": (
                            "你是污水智能化改造行业竞品情报分析师，负责从公开渠道整理竞品动态："
                            "新品发布、中标公告、技术动态、报价调整、客户案例。"
                            "每条动态尽量标注来源链接；无法确认则留空，禁止编造。只返回 JSON。"
                        ),
                    },
                    {"role": "user", "content": user_prompt},
                ],
                "temperature": 0.3,
                "search": ENABLE_DEEPSEEK_SEARCH,
            },
            timeout=LLM_TIMEOUT,
        )
        resp.raise_for_status()
        data = resp.json()
        content = (data.get("choices") or [{}])[0].get("message", {}).get("content", "").strip()
        total_tokens = (data.get("usage") or {}).get("total_tokens", "未知")
        parsed = _extract_json_block(content)
        if not parsed:
            logger.warning("[竞品AI] ❌ AI 返回内容无法解析为 JSON，竞品「%s」动态检索失败", comp.name)
            return None
        intel = _normalize_intel_payload(parsed)
        parsed["intel"] = intel

        intel_labels = [label for key, label in COMPETITOR_INTEL_CATEGORIES if intel.get(key)]
        logger.info(
            "[竞品AI] ✅ AI 检索成功"
            "\n  ├─ 竞品名称: %s"
            "\n  ├─ 消耗 tokens: %s"
            "\n  ├─ 检索到 %s 类动态: %s"
            "\n  └─ 信息源: %s",
            comp.name, total_tokens, len(intel_labels),
            "、".join(intel_labels) or "无新动态",
            intel.get("infoSource") or "未标注"
        )
        return parsed
    except Exception as exc:
        logger.warning("[竞品AI] ❌ 竞品「%s」动态检索异常（%s）", comp.name, exc)
        return None


def _apply_enrichment(comp: DBCompetitor, enriched: dict) -> List[str]:
    """仅在有有效新信息时更新字段（空字段不覆盖原有内容），返回被更新字段的中文名清单。"""
    changed: List[str] = []

    def _clean_str(val: Any) -> str:
        return val.strip() if isinstance(val, str) else ""

    def _clean_list(val: Any) -> List[str]:
        if isinstance(val, list):
            return [str(x).strip() for x in val if str(x).strip()]
        return []

    text_fields = [
        ("mainBusiness", "main_business", "主营业务"),
        ("techRoute", "tech_route", "技术路线"),
        ("pricingRange", "pricing_range", "报价水平"),
        ("typicalCases", "typical_cases", "典型案例"),
    ]
    for src_key, attr, label in text_fields:
        new_val = _clean_str(enriched.get(src_key))
        if new_val and new_val != (getattr(comp, attr) or ""):
            setattr(comp, attr, new_val)
            changed.append(label)

    for src_key, attr, label in [("strengths", "strengths", "优势分析"), ("weaknesses", "weaknesses", "劣势分析")]:
        new_list = _clean_list(enriched.get(src_key))
        if new_list and new_list != _safe_json_list(getattr(comp, attr)):
            setattr(comp, attr, json.dumps(new_list, ensure_ascii=False))
            changed.append(label)

    new_source = _clean_str(enriched.get("infoSource"))
    if new_source and new_source != (comp.info_source or ""):
        comp.info_source = new_source
        if "信息来源" not in changed:
            changed.append("信息来源")

    return changed


_CHANGE_LABEL_TO_KEY = {
    "主营业务": "mainBusiness",
    "技术路线": "techRoute",
    "报价水平": "pricingRange",
    "典型案例": "typicalCases",
    "优势分析": "strengths",
    "劣势分析": "weaknesses",
    "信息来源": "infoSource",
}


def _record_competitor_update(
    db: Session,
    comp: DBCompetitor,
    changed: List[str],
    enriched: dict,
    intel: Dict[str, Any],
) -> None:
    """将本次 AI 检索的动态与变更写入记录表，供详情页追溯。"""
    changed_map = {label: enriched.get(_CHANGE_LABEL_TO_KEY.get(label, ""), "") for label in changed}
    rec = DBCompetitorUpdate(
        id=str(uuid.uuid4()),
        competitor_id=comp.id,
        changed_fields=json.dumps(changed_map, ensure_ascii=False),
        intel_payload=json.dumps(intel, ensure_ascii=False),
        source=(intel.get("infoSource") or enriched.get("infoSource") or "DeepSeek 全网检索"),
        created_at=_now(),
    )
    db.add(rec)


@app.post("/api/competitors/{comp_id}/enrich")
def enrich_competitor(comp_id: str, payload: dict = None, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人", "商务专员"):
        raise HTTPException(status_code=403, detail="仅超级管理员、商务负责人或商务专员有权执行竞品智能补全")

    comp = db.query(DBCompetitor).filter(DBCompetitor.id == comp_id).first()
    if not comp:
        raise HTTPException(status_code=404, detail="竞品不存在")
    if not LLM_API_KEY:
        raise HTTPException(status_code=503, detail="请配置 LLM_API_KEY（DeepSeek）后使用竞品 AI 动态检索")

    logger.info("[竞品AI] 触发竞品智能补全 竞品=%s 操作者=%s(%s)", comp.name, current_user["name"], current_user["role"])
    enriched = _enrich_competitor_via_llm(comp)
    if not enriched:
        logger.warning("[竞品AI] ❌ 竞品「%s」智能补全失败（AI 未返回有效数据）", comp.name)
        raise HTTPException(status_code=502, detail="AI 检索失败，请稍后重试")

    intel = enriched.get("intel") or _normalize_intel_payload(enriched)
    if not _intel_has_content(intel):
        logger.info("[竞品AI] ⚠️ 竞品「%s」AI 检索未发现新的公开动态", comp.name)
        return {
            "competitor": competitor_to_api(comp),
            "intel": intel,
            "changedFields": [],
            "message": "未发现可更新 of 公开动态，已保留原有信息",
        }

    changed = _apply_enrichment(comp, enriched)
    comp.monitoring_intel = json.dumps(intel, ensure_ascii=False)
    comp.updated_at = _now()
    comp.last_enriched_at = _now()
    _record_competitor_update(db, comp, changed, enriched, intel)
    db.commit()
    db.refresh(comp)

    intel_labels = [label for key, label in COMPETITOR_INTEL_CATEGORIES if intel.get(key)]
    logger.info(
        "[竞品AI] ✅ 竞品「%s」动态已更新"
        "\n  ├─ 更新字段: %s"
        "\n  ├─ 动态类别: %s"
        "\n  └─ 操作者: %s",
        comp.name, "、".join(changed) or "无", "、".join(intel_labels) or "无",
        current_user["name"]
    )
    add_audit_log(
        db,
        current_user["name"],
        current_user["role"],
        "竞品动态检索",
        f"DeepSeek 检索竞品「{comp.name}」动态：{ '、'.join(intel_labels) or '已更新监测档案' }",
        comp.id,
    )
    return {
        "competitor": competitor_to_api(comp),
        "intel": intel,
        "changedFields": changed,
        "message": f"已检索到 {len(intel_labels)} 类动态：{'、'.join(intel_labels)}",
    }


@app.get("/api/competitors/{comp_id}/updates")
def list_competitor_updates(comp_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    rows = (
        db.query(DBCompetitorUpdate)
        .filter(DBCompetitorUpdate.competitor_id == comp_id)
        .order_by(DBCompetitorUpdate.created_at.desc())
        .all()
    )
    return [competitor_update_to_api(r) for r in rows]


@app.post("/api/competitors/enrich-all")
def enrich_all_competitors(payload: dict = None, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    role = current_user["role"]
    if role not in ("超级管理员", "商务负责人", "商务专员"):
        raise HTTPException(status_code=403, detail="仅超级管理员、商务负责人或商务专员有权执行一键全检索")

    if not LLM_API_KEY:
        raise HTTPException(status_code=503, detail="请配置 LLM_API_KEY（DeepSeek）后使用竞品智能补全")
    payload = payload or {}

    actives = [c for c in db.query(DBCompetitor).all() if (getattr(c, "status", "active") or "active") == "active"]
    logger.info("[竞品AI] ▶ 开始批量检索 %s 个活跃竞品 操作者=%s(%s)", len(actives), current_user["name"], current_user["role"])
    results = []
    for comp in actives:
        logger.info("[竞品AI] ─ 正在检索竞品「%s」（%s/%s）", comp.name, len(results) + 1, len(actives))
        enriched = _enrich_competitor_via_llm(comp)
        if not enriched:
            results.append({"id": comp.id, "name": comp.name, "changedFields": [], "message": "检索失败"})
            continue
        intel = enriched.get("intel") or _normalize_intel_payload(enriched)
        if not _intel_has_content(intel):
            results.append({"id": comp.id, "name": comp.name, "changedFields": [], "message": "无新动态"})
            continue
        changed = _apply_enrichment(comp, enriched)
        comp.monitoring_intel = json.dumps(intel, ensure_ascii=False)
        comp.updated_at = _now()
        comp.last_enriched_at = _now()
        _record_competitor_update(db, comp, changed, enriched, intel)
        db.commit()
        db.refresh(comp)
        add_audit_log(
            db,
            current_user["name"],
            current_user["role"],
            "竞品动态检索",
            f"批量检索竞品「{comp.name}」动态",
            comp.id,
        )
        results.append({"id": comp.id, "name": comp.name, "changedFields": changed, "message": "已更新"})
    updated_count = sum(1 for r in results if r.get("message") == "已更新")
    failed_count = sum(1 for r in results if r.get("message") == "检索失败")
    logger.info(
        "[竞品AI] ✅ 批量检索完成"
        "\n  ├─ 总计: %s 个"
        "\n  ├─ 已更新: %s 个"
        "\n  ├─ 检索失败: %s 个"
        "\n  └─ 无新动态: %s 个",
        len(actives), updated_count, failed_count,
        len(actives) - updated_count - failed_count
    )
    return {"total": len(actives), "updatedCount": updated_count, "results": results}


# ============ 定时竞品监测调度 ============

SCHEDULER_ENABLED = os.getenv("SCHEDULER_ENABLED", "true").strip().lower() == "true"
SCHEDULER_LOG = logging.getLogger("lead_backend.scheduler")


def run_scheduled_competitor_enrich() -> None:
    """定时执行竞品动态检索。

    调度策略（由 APScheduler 触发，本函数仅判断是否应执行）：
      - 重点竞品（priority=key）：距上次检索 >= 3 天
      - 普通竞品（priority=normal）：距上次检索 >= 7 天
      - 从未检索过的竞品：立即执行
    """
    db = SessionLocal()
    try:
        now = _now()
        key_interval = timedelta(days=3)
        normal_interval = timedelta(days=7)

        all_competitors = db.query(DBCompetitor).all()
        due_competitors: List[DBCompetitor] = []
        for comp in all_competitors:
            if (getattr(comp, "status", "active") or "active") != "active":
                continue
            last = getattr(comp, "last_enriched_at", None)
            priority = getattr(comp, "priority", None) or "normal"
            interval = key_interval if priority == "key" else normal_interval
            if last is None or (now - last) >= interval:
                due_competitors.append(comp)

        if not due_competitors:
            SCHEDULER_LOG.info("[调度] 本次无到期需检索的竞品")
            return

        SCHEDULER_LOG.info(
            "[调度] ▶ 开始定时检索 %s 个竞品（重点=%s 普通=%s）",
            len(due_competitors),
            sum(1 for c in due_competitors if (getattr(c, "priority", None) or "normal") == "key"),
            sum(1 for c in due_competitors if (getattr(c, "priority", None) or "normal") != "key"),
        )

        for comp in due_competitors:
            try:
                enriched = _enrich_competitor_via_llm(comp)
                if not enriched:
                    SCHEDULER_LOG.info("[调度] ─ 竞品「%s」检索无返回", comp.name)
                    continue
                intel = enriched.get("intel") or _normalize_intel_payload(enriched)
                if not _intel_has_content(intel):
                    SCHEDULER_LOG.info("[调度] ─ 竞品「%s」无新动态", comp.name)
                    continue
                changed = _apply_enrichment(comp, enriched)
                comp.monitoring_intel = json.dumps(intel, ensure_ascii=False)
                comp.updated_at = _now()
                comp.last_enriched_at = _now()
                _record_competitor_update(db, comp, changed, enriched, intel)
                intel_labels = [label for key, label in COMPETITOR_INTEL_CATEGORIES if intel.get(key)]
                SCHEDULER_LOG.info(
                    "[调度] ✅ 竞品「%s」已更新 动态=%s 字段=%s",
                    comp.name, "、".join(intel_labels) or "无", "、".join(changed) or "无"
                )
            except Exception as exc:
                SCHEDULER_LOG.error("[调度] ❌ 竞品「%s」检索异常: %s", comp.name, exc)
            db.commit()

        SCHEDULER_LOG.info("[调度] ✅ 定时检索完成，共处理 %s 个竞品", len(due_competitors))

        add_audit_log(
            db, "系统调度", "自动化服务", "竞品动态定时检索",
            f"定时任务完成竞品动态检索，处理 {len(due_competitors)} 个竞品"
        )
    except Exception as exc:
        SCHEDULER_LOG.error("[调度] ❌ 定时检索整体异常: %s", exc)
        db.rollback()
    finally:
        db.close()


from fastapi import UploadFile, File

LEAD_IMPORT_TEMPLATE_HEADERS = [
    "商务项目标题名称*", "业主/采购企业全名*", "行政所属区归口*",
    "细分行业归属", "物料分级归类", "招标企业指定联系人",
    "真实联系渠道电话", "采购诉求与痛点描述说明*",
]
LEAD_IMPORT_REQUIRED_FIELDS = ["商务项目标题名称*", "业主/采购企业全名*", "行政所属区归口*", "采购诉求与痛点描述说明*"]


def _score_dimensions(
    requirements: str,
    industry: str,
    region: str,
) -> Tuple[int, int, int]:
    """根据线索字段计算3维评分（导入用：匹配度、资质、区域）"""
    # match_score: 需求描述越详细分越高
    req_len = len(requirements or "")
    if req_len > 200:
        match_score = 85
    elif req_len > 100:
        match_score = 75
    elif req_len > 50:
        match_score = 65
    elif req_len > 10:
        match_score = 55
    else:
        match_score = 50

    # qualification_score: 行业匹配度
    industry_lower = (industry or "").lower()
    priority_keywords = ['水务', '水利', '环保', '能源', '供热', '电力', '监测', '监控', '自控', '自动化', '机电', '管网']
    matched = sum(1 for k in priority_keywords if k in industry_lower)
    if matched >= 2:
        qualification_score = 85
    elif matched >= 1:
        qualification_score = 70
    else:
        qualification_score = 50

    # region_score: 有明确地区加分
    if region and region != "未知":
        region_score = 70
    else:
        region_score = 50

    return match_score, qualification_score, region_score


def _build_lead_import_template() -> bytes:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "线索导入模板"
    ws.append(LEAD_IMPORT_TEMPLATE_HEADERS)
    widths = [35, 30, 15, 15, 15, 15, 18, 50]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.get("/api/leads/template")
def download_lead_import_template(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "查看角色":
        raise HTTPException(status_code=403, detail="查看角色无权下载导入模板")
    content = _build_lead_import_template()
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="lead_import_template.xlsx"'},
    )


@app.post("/api/leads/import")
def import_leads(
    file: UploadFile = File(...),
    duplicateAction: str = Form(default="skip"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] == "查看角色":
        raise HTTPException(status_code=403, detail="无权导入")

    try:
        from openpyxl import load_workbook
        file.file.seek(0)
        wb = load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法解析 Excel 文件: {str(e)}")

    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    if not rows:
        raise HTTPException(status_code=400, detail="Excel 文件中没有数据行")

    # 本次导入的统一时间戳（所有新线索共用，保证「新」标签正确）
    import_time_raw = _now()
    import_time_str = import_time_raw.strftime("%Y-%m-%d %H:%M:%S")

    # 校验列数
    expected_cols = len(LEAD_IMPORT_TEMPLATE_HEADERS)
    for row_idx, row in enumerate(rows, start=2):
        actual_cols = len(row)
        if actual_cols < expected_cols:
            raise HTTPException(
                status_code=400,
                detail=f"第{row_idx}行列数不足（需{expected_cols}列，实际{actual_cols}列），请使用最新导入模板",
            )
        break  # 只检查第一行

    duplicate_action = duplicateAction  # "skip" 或 "overwrite"
    scoring_rule = get_or_create_scoring_rule(db)  # noqa: F841

    imported = 0
    duplicated = 0
    overwritten = 0
    errors: list[str] = []

    for row_idx, row in enumerate(rows, start=2):
        try:
            title = str(row[0] or "").strip()
            company = str(row[1] or "").strip()
            if not title or not company:
                errors.append(f"第{row_idx}行: 商务项目标题名称和业主/采购企业全名不能为空")
                continue

            region = str(row[2] or "").strip() if len(row) > 2 else ""
            industry = str(row[3] or "").strip() if len(row) > 3 else ""
            category = str(row[4] or "").strip() if len(row) > 4 else ""
            contact_name = str(row[5] or "").strip() if len(row) > 5 else ""
            contact_phone = str(row[6] or "").strip() if len(row) > 6 else ""
            requirements = str(row[7] or "").strip() if len(row) > 7 else ""

            if not region:
                errors.append(f"第{row_idx}行: 行政所属区归口不能为空")
                continue
            if not requirements:
                errors.append(f"第{row_idx}行: 采购诉求与痛点描述说明不能为空")
                continue

            # 去重检查
            existing = db.query(DBLead).filter(
                DBLead.title == title,
                DBLead.company == company,
            ).first()

            if existing:
                if duplicate_action == "skip":
                    duplicated += 1
                    continue
                # overwrite: 更新已有记录
                existing.contact_name = contact_name or existing.contact_name
                existing.contact_phone = contact_phone or existing.contact_phone
                existing.requirements = requirements or existing.requirements
                existing.region = region or existing.region
                existing.industry = industry or existing.industry
                existing.category = category or existing.category
                # 重新打分
                scores = _score_dimensions(existing.requirements or "", existing.industry or "", existing.region or "")
                existing.match_score, existing.qualification_score, existing.region_score = scores
                total, grade = calculate_score(
                    scoring_rule,
                    existing.budget_score or 50,
                    existing.match_score,
                    50,
                    existing.qualification_score,
                    existing.region_score,
                )
                existing.total_score = Decimal(str(total))
                existing.grade = grade
                existing.crawl_time = import_time_raw
                db.commit()
                overwritten += 1
                add_audit_log(
                    db, current_user["name"], current_user["role"],
                    "线索批量导入-覆盖", f"覆盖已有线索「{title}」", existing.id,
                )
                continue

            # 新建线索
            lead_id = str(uuid.uuid4())
            # 3维打分（match, qualification, region），budget_score/stage_score 默认50
            scores = _score_dimensions(requirements, industry, region)
            match_score, qualification_score, region_score = scores
            total, grade = calculate_score(scoring_rule, 50, match_score, 50, qualification_score, region_score)
            new_lead = DBLead(
                id=lead_id,
                title=title,
                company=company,
                publish_date=date.today(),
                region=region or "未知",
                industry=industry or "其他",
                category=category or "",
                requirements=requirements,
                contact_name=contact_name,
                contact_phone=contact_phone,
                status="待分配",
                crawl_time=import_time_raw,
                lead_source="import",
                created_at=import_time_raw,
                budget_score=50,
                match_score=match_score,
                stage_score=50,
                qualification_score=qualification_score,
                region_score=region_score,
                total_score=Decimal(str(total)),
                grade=grade,
                created_by=current_user["name"],
            )
            db.add(new_lead)
            db.commit()
            imported += 1
            add_audit_log(
                db, current_user["name"], current_user["role"],
                "线索批量导入", f"导入新线索「{title}」", lead_id,
            )
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    return {"imported": imported, "duplicated": duplicated, "overwritten": overwritten, "errors": errors, "importTime": import_time_str}


@app.get("/api/leads/scheduled-result")
def get_scheduled_crawl_result():
    return _scheduled_crawl_result


# 最近一次定时爬取结果（用于前端轮询提醒）
_scheduled_crawl_result: dict = {"timestamp": "", "summary": ""}


def run_scheduled_crawling() -> None:
    db = SessionLocal()
    global _scheduled_crawl_result
    try:
        configs = db.query(DBSearchConfig).all()
        if not configs:
            SCHEDULER_LOG.info("[调度] 无网监配置，跳过商机自动爬取")
            return
        SCHEDULER_LOG.info(f"[调度] 开始执行商机自动爬取，当前有 {len(configs)} 个配置规则")
        db_config = search_config_to_api(configs[0])
        search_rules = normalize_search_config(db_config)
        run_id = str(uuid.uuid4())[:8]
        candidates, scan_stats = build_crawler_candidates(search_rules, run_id, db=db)
        rules = get_or_create_scoring_rule(db)
        integrated_leads, duplicate_count = attach_scores_and_stage_leads(db, candidates, rules, run_id)
        db.commit()
        msg = f"新增 {len(integrated_leads)} 条，过滤 {scan_stats.get('filtered', 0)} 条，重复 {duplicate_count} 条"
        SCHEDULER_LOG.info("[调度] 商机爬取完成 run_id=%s %s", run_id, msg)
        _scheduled_crawl_result = {
            "timestamp": _now().strftime("%Y-%m-%d %H:%M:%S"),
            "summary": msg,
        }
    except Exception as exc:
        db.rollback()
        SCHEDULER_LOG.error(f"[调度] 商机爬取调度异常: {exc}")
    finally:
        db.close()

def start_scheduler() -> None:
    """启动 APScheduler 定时监测。"""
    if not SCHEDULER_ENABLED:
        SCHEDULER_LOG.info("[调度] SCHEDULER_ENABLED=false，跳过定时任务")
        return

    from apscheduler.schedulers.background import BackgroundScheduler

    scheduler = BackgroundScheduler()
    scheduler.add_job(
        run_scheduled_competitor_enrich,
        trigger="interval",
        hours=3,
        id="competitor_enrich",
        name="竞品动态监测（每3小时判断到期竞品）",
        misfire_grace_time=3600,
    )
    # 每日早8:00执行全量爬取
    scheduler.add_job(
        run_scheduled_crawling,
        trigger="cron",
        hour=8,
        minute=0,
        id="leads_crawl_daily",
        name="公域商机全量爬取（每日8点）",
    )
    # 每4小时增量爬取高优先级
    scheduler.add_job(
        run_scheduled_crawling,
        trigger="interval",
        hours=4,
        id="leads_crawl_incremental",
        name="公域商机高优增量爬取（每4小时）",
    )
    scheduler.start()
    SCHEDULER_LOG.info(
        "[调度] ✅ 定时任务调度已全量启动（包含竞品监测、商机爬取等）"
    )


# 服务启动时初始化调度
start_scheduler()


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)

